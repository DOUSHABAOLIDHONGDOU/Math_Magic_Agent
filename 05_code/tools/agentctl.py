#!/usr/bin/env python3
"""Small workflow helper for the Math Magic multi-agent project."""

from __future__ import annotations

import argparse
import contextlib
import csv
import datetime as dt
import importlib.util
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import time
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[2]
STATE_PATH = PROJECT_ROOT / "00_shared" / "workflow_state.json"
LOCK_PATH = PROJECT_ROOT / "00_shared" / ".workflow_state.lock"
QUESTIONS = ["Q1", "Q2", "Q3", "Q4", "Q5"]
SCHEMES = ["A", "B", "C"]
STAGES = [
    "INIT",
    "PROBLEM_LOADED",
    "SCHEMES_GENERATED",
    "SCHEMES_APPROVED",
    "CLAUDE_WORKORDERS_CREATED",
    "CODE_COMPLETED",
    "CODE_REVIEWED",
    "MODEL_CONFIRMED",
    "FIGURES_GENERATED",
    "FIGURES_APPROVED",
    "PAPER_WRITTEN",
    "APPENDIX_READY",
    "FINAL_REVIEW",
]
SELECTED_SCHEME_STATUSES = {
    "approved_to_run",
    "workorder_created",
    "code_completed",
    "review_template_created",
    "review_pass",
    "review_revise",
    "review_blocked",
}


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def today() -> str:
    return dt.date.today().isoformat()


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def default_state() -> dict:
    return {
        "version": "0.2.24",
        "stage": "INIT",
        "current_question": None,
        "workflow": {
            "mode": "sequential",
            "allow_parallel_questions": False,
            "question_dependency_rule": "later questions wait until previous question model_confirmed",
        },
        "language": {
            "recommended": "Python",
            "approved": False,
            "decision_id": None,
        },
        "problem": {
            "title": None,
            "statement_file": "01_problem/problem_statement.md",
            "data_dictionary": "01_problem/data_dictionary.md",
            "raw_data_dir": None,
            "question_ids": [],
        },
        "questions": {
            q: {
                "status": "not_started",
                "schemes_generated": False,
                "schemes_approved": False,
                "scheme_decision_id": None,
                "workorders_created": False,
                "code_completed": False,
                "code_reviewed": False,
                "model_confirmed": False,
                "confirmed_scheme": None,
                "model_decision_id": None,
                "figures_generated": False,
                "figures_approved": False,
                "figure_decision_id": None,
                "figure_language": None,
                "paper_written": False,
                "paper_section": None,
                "latex_compiled_at": None,
                "schemes": {
                    s: {
                        "status": "not_started",
                        "workorder": None,
                        "claude_prompt": None,
                        "dispatch_status": None,
                        "dispatch_mode": None,
                        "dispatch_prompt": None,
                        "dispatch_log": None,
                        "dispatch_terminal_script": None,
                        "dispatch_terminal_status": None,
                        "completion_report": None,
                        "review": None,
                    }
                    for s in SCHEMES
                },
            }
            for q in QUESTIONS
        },
        "artifacts": [],
        "updated_at": now_iso(),
    }


def now_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat()


def load_state() -> dict:
    if not STATE_PATH.exists():
        state = default_state()
        save_state(state)
        return state
    state = json.loads(read_text(STATE_PATH))
    changed = migrate_state(state)
    if changed:
        save_state(state)
    return state


def migrate_state(state: dict) -> bool:
    changed = False
    default = default_state()
    if state.get("version") != default["version"]:
        state["version"] = default["version"]
        changed = True
    if "workflow" not in state:
        state["workflow"] = default["workflow"]
        changed = True
    if "question_ids" not in state.get("problem", {}):
        state.setdefault("problem", {})["question_ids"] = []
        changed = True
    state.setdefault("questions", {})
    for question in QUESTIONS:
        if question not in state["questions"]:
            state["questions"][question] = default["questions"][question]
            changed = True
            continue
        qstate = state["questions"][question]
        default_qstate = default["questions"][question]
        for key, value in default_qstate.items():
            if key not in qstate:
                qstate[key] = value
                changed = True
        qstate.setdefault("schemes", {})
        for scheme in SCHEMES:
            if scheme not in qstate["schemes"]:
                qstate["schemes"][scheme] = default_qstate["schemes"][scheme]
                changed = True
                continue
            for key, value in default_qstate["schemes"][scheme].items():
                if key not in qstate["schemes"][scheme]:
                    qstate["schemes"][scheme][key] = value
                    changed = True
    return changed


def save_state(state: dict) -> None:
    state["updated_at"] = now_iso()
    write_text(STATE_PATH, json.dumps(state, ensure_ascii=False, indent=2) + "\n")


@contextlib.contextmanager
def workflow_lock():
    LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LOCK_PATH.open("w", encoding="utf-8") as lock_file:
        try:
            import fcntl

            fcntl.flock(lock_file, fcntl.LOCK_EX)
        except ImportError:
            fcntl = None
        try:
            yield
        finally:
            if "fcntl" in locals() and fcntl is not None:
                fcntl.flock(lock_file, fcntl.LOCK_UN)


def set_stage(state: dict, stage: str) -> None:
    if stage not in STAGES:
        raise SystemExit(f"unknown stage: {stage}")
    state["stage"] = stage


def append_artifact(state: dict, kind: str, path: Path, note: str = "") -> None:
    artifact_path = rel(path)
    state.setdefault("artifacts", [])
    already_recorded = any(
        item.get("kind") == kind and item.get("path") == artifact_path and item.get("note", "") == note
        for item in state["artifacts"]
    )
    if not already_recorded:
        state["artifacts"].append({"kind": kind, "path": artifact_path, "note": note, "created_at": now_iso()})


def expected_completion_path(question: str, scheme: str) -> Path:
    return PROJECT_ROOT / "04_claude_workorders" / "completions" / f"{question}_scheme_{scheme}_completion.md"


def expected_standard_outputs(question: str, scheme: str) -> list[Path]:
    return [
        PROJECT_ROOT / "06_results" / question / "tables" / f"scheme_{scheme}_metrics.csv",
        PROJECT_ROOT / "06_results" / question / "figures" / f"scheme_{scheme}_raw.png",
        PROJECT_ROOT / "06_results" / question / "logs" / f"scheme_{scheme}_run.md",
    ]


def question_section_path(question: str) -> Path:
    return PROJECT_ROOT / "07_paper" / "sections" / f"model_{question.lower()}.tex"


def paper_summary_sections() -> list[Path]:
    return [
        PROJECT_ROOT / "07_paper" / "sections" / "abstract.tex",
        PROJECT_ROOT / "07_paper" / "sections" / "problem_analysis.tex",
        PROJECT_ROOT / "07_paper" / "sections" / "model_validation.tex",
        PROJECT_ROOT / "07_paper" / "sections" / "evaluation.tex",
    ]


def dispatch_log_dir() -> Path:
    return PROJECT_ROOT / "04_claude_workorders" / "dispatch_logs"


def dispatch_terminal_dir() -> Path:
    return PROJECT_ROOT / "04_claude_workorders" / "terminal_runs"


def dispatch_monitor_dir() -> Path:
    return dispatch_terminal_dir() / "monitors"


def vscode_tasks_path() -> Path:
    return PROJECT_ROOT / ".vscode" / "tasks.json"


def current_target_os() -> str:
    return "windows" if sys.platform.startswith("win") else "posix"


def ensure_target_os(value: str) -> str:
    value = (value or "auto").lower()
    if value == "auto":
        return current_target_os()
    if value in {"windows", "posix"}:
        return value
    raise SystemExit("target os must be auto, windows, or posix")


def ps_single_quote(value: str | Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def dispatch_terminal_status_path() -> Path:
    return dispatch_terminal_dir() / "CURRENT_TERMINAL_STATUS.json"


def dispatch_config_path() -> Path:
    return PROJECT_ROOT / "04_claude_workorders" / "claude_dispatch_config.json"


def load_dispatch_config() -> dict:
    path = dispatch_config_path()
    if path.exists():
        return json.loads(read_text(path))
    return {}


def format_claude_command(executable: str | Path) -> str:
    return f"{shlex.quote(str(executable))} -p --permission-mode acceptEdits"


def discover_claude_binary() -> Path | str | None:
    for command in ["claude", "claude-code"]:
        executable = shutil.which(command)
        if executable:
            return command
    extension_roots = [
        Path.home() / ".vscode" / "extensions",
        Path.home() / ".vscode-insiders" / "extensions",
        Path.home() / ".cursor" / "extensions",
        Path.home() / ".windsurf" / "extensions",
    ]
    candidates: list[Path] = []
    for root in extension_roots:
        if not root.exists():
            continue
        candidates.extend(root.glob("anthropic.claude-code-*/resources/native-binary/claude"))
    candidates = [path for path in candidates if path.exists() and os.access(path, os.X_OK)]
    if not candidates:
        return None
    return sorted(candidates, key=lambda path: path.stat().st_mtime)[-1]


def auto_discovered_claude_command() -> str:
    binary = discover_claude_binary()
    if not binary:
        return ""
    return format_claude_command(binary)


def configured_claude_command(args: argparse.Namespace) -> str:
    if getattr(args, "command", ""):
        return args.command
    env_command = os.environ.get("CLAUDE_CODE_COMMAND", "")
    if env_command:
        return env_command
    config = load_dispatch_config()
    if config.get("command", ""):
        return config["command"]
    if config.get("auto_discover", True):
        return auto_discovered_claude_command()
    return ""


def latest_revision_prompt(question: str, scheme: str) -> Path | None:
    pattern = f"{question}_scheme_{scheme}_revision_prompt_*.md"
    matches = sorted((PROJECT_ROOT / "04_claude_workorders").glob(pattern), key=lambda p: p.stat().st_mtime)
    return matches[-1] if matches else None


def resolve_dispatch_prompt(state: dict, question: str, scheme: str, prompt: Path | None, use_revision: bool) -> Path:
    if prompt:
        path = prompt if prompt.is_absolute() else PROJECT_ROOT / prompt
        if not path.exists():
            raise SystemExit(f"prompt not found: {path}")
        return path
    if use_revision:
        revision = latest_revision_prompt(question, scheme)
        if revision:
            return revision
        raise SystemExit(f"no revision prompt found for {question} scheme {scheme}")
    scheme_state = state["questions"][question]["schemes"][scheme]
    if scheme_state.get("claude_prompt"):
        path = PROJECT_ROOT / scheme_state["claude_prompt"]
        if path.exists():
            return path
    default_path = PROJECT_ROOT / "04_claude_workorders" / f"{question}_scheme_{scheme}_claude_prompt.md"
    if default_path.exists():
        return default_path
    raise SystemExit(f"no Claude prompt found for {question} scheme {scheme}")


def record_claude_completion(state: dict, question: str, scheme: str, report_path: Path) -> Path:
    dest = expected_completion_path(question, scheme)
    if report_path.resolve() != dest.resolve():
        write_text(dest, read_text(report_path))
    qstate = state["questions"][question]
    scheme_state = qstate["schemes"][scheme]
    scheme_state["completion_report"] = rel(dest)
    scheme_state["status"] = "code_completed"
    scheme_state["detected_at"] = now_iso()
    selected = approved_schemes(qstate) or [scheme]
    qstate["code_completed"] = all(qstate["schemes"][s]["completion_report"] for s in selected)
    if qstate["code_completed"]:
        qstate["status"] = "code_completed"
        set_stage(state, "CODE_COMPLETED")
    append_artifact(state, "claude_completion_report", dest, f"{question} scheme {scheme}")
    return dest


def create_review_file(state: dict, question: str, scheme: str) -> Path:
    out_path = PROJECT_ROOT / "06_results" / question / "logs" / f"scheme_{scheme}_codex_review.md"
    report = state["questions"][question]["schemes"][scheme].get("completion_report") or "待 Claude Code 提供"
    text = render_codex_review_template(question, scheme, report)
    write_text(out_path, text)
    state["questions"][question]["schemes"][scheme]["review"] = rel(out_path)
    state["questions"][question]["schemes"][scheme]["status"] = "review_template_created"
    append_artifact(state, "codex_review", out_path, f"{question} scheme {scheme}")
    return out_path


def append_markdown_log(path: Path, title: str, lines: list[str]) -> None:
    existing = read_text(path) if path.exists() else f"# {path.stem}\n"
    block = "\n\n" + f"## {title}\n\n" + "\n".join(lines).rstrip() + "\n"
    write_text(path, existing.rstrip() + block)


def print_user_facing_brief(path: Path, text: str, brief_type: str) -> None:
    print(f"USER_FACING_BRIEF: {brief_type}")
    print(f"path: {rel(path)}")
    print()
    print(text.rstrip())
    print()
    print("ACTION_REQUIRED: Codex must paste/summarize these options in the user chat and wait for approval.")


def ensure_question(question: str) -> str:
    question = question.upper()
    if question not in QUESTIONS:
        raise SystemExit(f"question must be one of {', '.join(QUESTIONS)}")
    return question


def question_choices() -> list[str]:
    return QUESTIONS + [q.lower() for q in QUESTIONS]


def ensure_scheme(scheme: str) -> str:
    scheme = scheme.upper()
    if scheme not in SCHEMES:
        raise SystemExit(f"scheme must be one of {', '.join(SCHEMES)}")
    return scheme


def active_question_ids(state: dict) -> list[str]:
    return state["problem"].get("question_ids") or QUESTIONS


def next_question_to_solve(state: dict) -> str | None:
    for question in active_question_ids(state):
        if not state["questions"][question].get("model_confirmed"):
            return question
    return None


def previous_question(state: dict, question: str) -> str | None:
    questions = active_question_ids(state)
    if question not in questions:
        return None
    index = questions.index(question)
    return questions[index - 1] if index > 0 else None


def assert_question_unlocked(state: dict, question: str, force: bool = False) -> None:
    if force or state.get("workflow", {}).get("allow_parallel_questions"):
        return
    if question not in active_question_ids(state):
        raise SystemExit(f"{question} is not an active question for the imported problem")
    next_question = next_question_to_solve(state)
    if next_question and question != next_question:
        prev = previous_question(state, question)
        reason = f"{prev} has not been model_confirmed" if prev else f"current question is {next_question}"
        raise SystemExit(
            f"sequential workflow guard: {question} is locked; solve {next_question} first ({reason}). "
            "Use --force only for explicit diagnostic tests."
        )


def approved_schemes(qstate: dict) -> list[str]:
    return [scheme for scheme in SCHEMES if qstate["schemes"][scheme].get("status") in SELECTED_SCHEME_STATUSES]


def command_status(args: argparse.Namespace) -> None:
    state = load_state()
    state_path = PROJECT_ROOT / "00_shared" / "PROJECT_STATE.md"
    boundary_path = PROJECT_ROOT / "00_shared" / "QUESTION_BOUNDARIES.md"
    print(f"Project root: {PROJECT_ROOT}")
    print(f"Machine state: {STATE_PATH}")
    print(f"Stage: {state['stage']}")
    print(f"Updated at: {state['updated_at']}")
    print()
    print("== Project State ==")
    print(extract_section(read_text(state_path), "当前阶段"))
    print(extract_section(read_text(state_path), "全局技术路线"))
    print(extract_section(read_text(state_path), "三次审批状态"))
    print()
    print("== Boundaries ==")
    print(extract_section(read_text(boundary_path), "待确认边界"))


def command_init_state(args: argparse.Namespace) -> None:
    if STATE_PATH.exists() and not args.force:
        print(STATE_PATH)
        print("state already exists; use --force to reset")
        return
    state = default_state()
    save_state(state)
    print(STATE_PATH)


def command_import_problem(args: argparse.Namespace) -> None:
    state = load_state()
    title = args.title or "题目名称待定"
    statement_text = ""
    if args.statement:
        statement_text = read_text(Path(args.statement))
    question_ids = detect_question_ids(statement_text)
    if not question_ids:
        question_ids = QUESTIONS[: args.num_questions]
    problem_path = PROJECT_ROOT / "01_problem" / "problem_statement.md"
    question_rows = [
        f"| {qid} | 待 Codex 拆解 | 待填写 | 待填写 | 待填写 |" for qid in question_ids
    ]
    body = [
        "# Problem Statement",
        "",
        "## 题目信息",
        "",
        "- 竞赛类型：赛前训练",
        f"- 题目编号：{args.problem_id or '待填写'}",
        f"- 题目名称：{title}",
        "- 附件列表：待数据扫描后补充",
        "",
        "## 原始题面",
        "",
        statement_text or "待粘贴或导入。",
        "",
        "## 问题拆解",
        "",
        "| 问题 | 原题要求 | 实际建模任务 | 必须输出 | 可选输出 |",
        "|---|---|---|---|---|",
        *question_rows,
        "",
        "## 隐含约束",
        "",
        "- 待 Codex 根据题面整理。",
        "",
        "## 评价指标",
        "",
        "- 待 Codex 根据题面整理。",
        "",
    ]
    write_text(problem_path, "\n".join(body))
    state["problem"]["title"] = title
    state["problem"]["statement_file"] = rel(problem_path)
    state["problem"]["question_ids"] = question_ids
    state["current_question"] = question_ids[0] if question_ids else None
    if args.data_dir:
        state["problem"]["raw_data_dir"] = rel(Path(args.data_dir))
    set_stage(state, "PROBLEM_LOADED")
    append_artifact(state, "problem_statement", problem_path, "imported problem statement")
    save_state(state)
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} 题目导入",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl import-problem",
            f"- 用途：导入训练题目 `{title}`。",
            f"- 关联文件：`{rel(problem_path)}`",
        ],
    )
    print(problem_path)


def detect_question_ids(text: str) -> list[str]:
    found = []
    for match in re.finditer(r"问题\s*([1-9])", text):
        qid = f"Q{match.group(1)}"
        if qid in QUESTIONS and qid not in found:
            found.append(qid)
    return found


def command_scan_data(args: argparse.Namespace) -> None:
    state = load_state()
    data_dir = Path(args.data_dir)
    if not data_dir.is_absolute():
        data_dir = PROJECT_ROOT / data_dir
    if not data_dir.exists():
        raise SystemExit(f"data dir not found: {data_dir}")
    rows = []
    for path in sorted(p for p in data_dir.rglob("*") if p.is_file()):
        rows.extend(inspect_data_file(path, sample_rows=args.sample_rows, include_unsupported=args.include_unsupported))
    out_path = Path(args.out) if args.out else PROJECT_ROOT / "01_problem" / "data_dictionary.md"
    if not out_path.is_absolute():
        out_path = PROJECT_ROOT / out_path
    write_data_dictionary(out_path, rows, data_dir)
    state["problem"]["raw_data_dir"] = rel(data_dir)
    state["problem"]["data_dictionary"] = rel(out_path)
    append_artifact(state, "data_dictionary", out_path, "scanned raw data files")
    save_state(state)
    print(out_path)


def inspect_data_file(path: Path, sample_rows: int = 200, include_unsupported: bool = False) -> list[dict]:
    suffix = path.suffix.lower()
    rows: list[dict] = []
    try:
        if suffix in [".csv", ".txt"]:
            import pandas as pd

            df = pd.read_csv(path, nrows=sample_rows)
            total_rows = count_text_rows(path)
            rows.extend(column_rows(path, df, total_rows=total_rows, sheet=""))
        elif suffix in [".xlsx", ".xls"]:
            import pandas as pd

            xls = pd.ExcelFile(path)
            excel_rows = count_excel_rows(path) if suffix == ".xlsx" else {}
            for sheet in xls.sheet_names:
                df = pd.read_excel(path, sheet_name=sheet, nrows=sample_rows)
                rows.extend(column_rows(path, df, total_rows=excel_rows.get(sheet), sheet=sheet))
        elif suffix == ".json":
            import pandas as pd

            try:
                df = pd.read_json(path, lines=True, nrows=sample_rows)
            except TypeError:
                df = pd.read_json(path)
            rows.extend(column_rows(path, df.head(sample_rows), total_rows=None, sheet=""))
        elif include_unsupported:
            rows.append(file_row(path, "", "", "", f"unsupported file type {suffix}"))
    except Exception as exc:  # noqa: BLE001 - this is a diagnostic tool.
        rows.append(file_row(path, "", "", "", f"scan failed: {exc}"))
    return rows


def count_text_rows(path: Path) -> int | None:
    try:
        with path.open("rb") as f:
            return max(sum(1 for _ in f) - 1, 0)
    except OSError:
        return None


def count_excel_rows(path: Path) -> dict[str, int]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        return {sheet.title: max((sheet.max_row or 1) - 1, 0) for sheet in workbook.worksheets}
    except Exception:  # noqa: BLE001 - row counts are helpful but noncritical.
        return {}


def column_rows(path: Path, df, total_rows: int | None, sheet: str) -> list[dict]:
    rows = []
    for col in df.columns:
        series = df[col]
        rows.append(
            file_row(
                path,
                sheet,
                str(col),
                str(series.dtype),
                f"sample_non_null={int(series.notna().sum())}; total_rows={total_rows if total_rows is not None else 'unknown'}",
            )
        )
    if len(df.columns) == 0:
        rows.append(file_row(path, sheet, "", "", "empty or no columns detected"))
    return rows


def file_row(path: Path, sheet: str, column: str, dtype: str, note: str) -> dict:
    return {
        "file": rel(path),
        "sheet": sheet,
        "column": column,
        "dtype": dtype,
        "unit": "待 Codex 判断",
        "meaning": "待 Codex 根据题面判断",
        "quality": note,
    }


def write_data_dictionary(out_path: Path, rows: list[dict], data_dir: Path) -> None:
    lines = [
        "# Data Dictionary",
        "",
        f"- 数据目录：`{rel(data_dir)}`",
        f"- 扫描时间：{now_iso()}",
        "",
        "## 字段字典",
        "",
        "| 文件 | 工作表 | 字段 | 类型 | 单位 | 含义 | 数据质量 |",
        "|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        lines.append(
            f"| `{row['file']}` | {row['sheet'] or '-'} | {row['column'] or '-'} | {row['dtype'] or '-'} | {row['unit']} | {row['meaning']} | {row['quality']} |"
        )
    lines.extend(
        [
            "",
            "## 数据清洗规则",
            "",
            "- 待 Claude Code 根据 Codex 工单实现，Codex 负责审查。",
            "",
            "## 数据使用风险",
            "",
            "- 待 Codex 根据题面、字段和缺失异常情况判断。",
            "",
        ]
    )
    write_text(out_path, "\n".join(lines))


def extract_section(text: str, title: str) -> str:
    marker = f"## {title}"
    start = text.find(marker)
    if start == -1:
        return f"{marker}\n未找到。"
    next_start = text.find("\n## ", start + len(marker))
    if next_start == -1:
        next_start = len(text)
    return text[start:next_start].strip()


def command_env_check(args: argparse.Namespace) -> None:
    modules = [
        "pypdf",
        "fitz",
        "pytesseract",
        "numpy",
        "pandas",
        "matplotlib",
        "seaborn",
        "scipy",
        "sklearn",
        "openpyxl",
        "xlrd",
        "statsmodels",
        "networkx",
    ]
    print("== Python Modules ==")
    for module in modules:
        ok = importlib.util.find_spec(module) is not None
        print(f"{module}: {'ok' if ok else 'missing'}")
    print()
    print("== Commands ==")
    for command in ["xelatex", "tesseract"]:
        found = shutil.which(command)
        print(f"{command}: {found or 'missing'}")
    tesseract = shutil.which("tesseract")
    if tesseract:
        result = subprocess.run(
            [tesseract, "--list-langs"],
            text=True,
            capture_output=True,
            check=False,
        )
        langs = result.stdout
        print(f"tesseract lang chi_sim: {'ok' if 'chi_sim' in langs else 'missing'}")
        print(f"tesseract lang eng: {'ok' if 'eng' in langs else 'missing'}")
    print()
    cache_dir = PROJECT_ROOT / ".cache" / "matplotlib"
    cache_dir.mkdir(parents=True, exist_ok=True)
    print(f"Matplotlib cache: {cache_dir}")


def command_doctor(args: argparse.Namespace) -> None:
    target_os = ensure_target_os(args.target_os)
    modules = [
        "pypdf",
        "fitz",
        "pytesseract",
        "numpy",
        "pandas",
        "matplotlib",
        "seaborn",
        "scipy",
        "sklearn",
        "openpyxl",
        "xlrd",
        "statsmodels",
        "networkx",
        "PIL",
    ]
    commands = ["python", "xelatex", "tesseract", "node", "npm", "claude", "code"]
    checks: list[tuple[str, bool, str]] = []
    for module in modules:
        checks.append((f"python module {module}", importlib.util.find_spec(module) is not None, "install environment.yml"))
    for command in commands:
        found = shutil.which(command)
        required = command != "code"
        checks.append((f"command {command}", bool(found) or not required, found or ("optional but useful for VS Code tasks" if command == "code" else "missing")))
    excellent_dir = PROJECT_ROOT / "02_references" / "excellent_papers"
    excellent_count = len(list(excellent_dir.rglob("*.pdf"))) if excellent_dir.exists() else 0
    checks.append(("excellent papers", excellent_count > 0, f"{excellent_count} pdf files"))
    checks.append(("paper template", (PROJECT_ROOT / "07_paper" / "main.tex").exists(), "07_paper/main.tex"))
    checks.append(("template raw archive", (PROJECT_ROOT / "07_paper" / "template_raw").exists(), "07_paper/template_raw"))
    checks.append(("workflow state", STATE_PATH.exists(), rel(STATE_PATH)))
    if args.write_vscode_smoke_task:
        task_path = write_vscode_smoke_task(target_os=target_os)
        checks.append(("VS Code Claude smoke task", True, rel(task_path)))
    else:
        checks.append(("VS Code tasks", vscode_tasks_path().exists(), rel(vscode_tasks_path())))
    print("== Math Magic Doctor ==")
    print(f"target_os: {target_os}")
    failed = False
    for name, ok, detail in checks:
        print(f"{name}: {'ok' if ok else 'missing'} ({detail})")
        failed = failed or not ok
    print()
    if args.write_vscode_smoke_task:
        print("Run VS Code task: Math Magic: Claude smoke test")
        print("It should open an integrated terminal and print the Claude Code version.")
    if failed and args.strict:
        raise SystemExit(1)


def write_vscode_smoke_task(target_os: str) -> Path:
    path = vscode_tasks_path()
    if path.exists():
        try:
            payload = json.loads(read_text(path))
        except json.JSONDecodeError as exc:
            raise SystemExit(f"cannot update invalid VS Code tasks JSON: {path}: {exc}") from exc
    else:
        payload = {"version": "2.0.0", "tasks": []}
    payload["version"] = payload.get("version") or "2.0.0"
    tasks = payload.setdefault("tasks", [])
    smoke_label = "Math Magic: Claude smoke test"
    tasks[:] = [task for task in tasks if task.get("label") != smoke_label]
    if target_os == "windows":
        command = "powershell"
        args = [
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            "Write-Host 'Math Magic Claude smoke test'; claude --version; python 05_code/tools/agentctl.py env-check",
        ]
    else:
        command = "bash"
        args = ["-lc", "echo 'Math Magic Claude smoke test'; claude --version; python 05_code/tools/agentctl.py env-check"]
    tasks.append(vscode_shell_task(smoke_label, command, args, dedicated_panel=True, focus=True))
    write_text(path, json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    return path


def command_tools(args: argparse.Namespace) -> None:
    registry_path = PROJECT_ROOT / "05_code" / "tools" / "tool_registry.json"
    registry = json.loads(read_text(registry_path))
    print(f"Tool registry version: {registry['version']}")
    print(f"Project: {registry['project']}")
    print()
    for tool in registry["tools"]:
        print(f"== {tool['id']} ==")
        print(f"name: {tool['name']}")
        print(f"path: {tool['path']}")
        print(f"owner: {tool['owner']}")
        print(f"purpose: {tool['purpose']}")
        print("commands:")
        for command in tool.get("command_examples", []):
            print(f"  - {command}")
        print()


def command_create_workorder(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    assert_question_unlocked(state, question, force=args.force)
    if state["questions"][question]["schemes"][scheme].get("status") not in SELECTED_SCHEME_STATUSES and not args.force:
        raise SystemExit(f"{question} scheme {scheme} is not approved; run approve-schemes first")
    out_path = create_workorder(
        question=question,
        scheme=scheme,
        workorder_id=args.workorder_id,
        out=Path(args.out) if args.out else None,
    )
    state["questions"][question]["schemes"][scheme]["workorder"] = rel(out_path)
    append_artifact(state, "claude_workorder", out_path, f"{question} scheme {scheme}")
    save_state(state)
    print(out_path)


def create_workorder(question: str, scheme: str, workorder_id: str | None = None, out: Path | None = None) -> Path:
    template_path = PROJECT_ROOT / "04_claude_workorders" / "templates" / "claude_workorder_template.md"
    template = read_text(template_path)
    workorder_id = workorder_id or f"{question}_{scheme}_001"
    out_path = out or PROJECT_ROOT / "04_claude_workorders" / f"{question}_scheme_{scheme}_workorder_001.md"
    if not out_path.is_absolute():
        out_path = PROJECT_ROOT / out_path
    text = template
    text = text.replace("- 工单 ID：", f"- 工单 ID：{workorder_id}")
    text = text.replace("- 相关问题：Q1 / Q2 / Q3", f"- 相关问题：{question}")
    text = text.replace("- 相关问题：QX", f"- 相关问题：{question}")
    text = text.replace("- 方案：A / B / C", f"- 方案：{scheme}")
    text = text.replace("- 方案：X", f"- 方案：{scheme}")
    text = text.replace("`06_results/QX/tables/scheme_X_metrics.csv`", f"`06_results/{question}/tables/scheme_{scheme}_metrics.csv`")
    text = text.replace("`06_results/QX/figures/scheme_X_raw.png`", f"`06_results/{question}/figures/scheme_{scheme}_raw.png`")
    text = text.replace("`06_results/QX/logs/scheme_X_run.md`", f"`06_results/{question}/logs/scheme_{scheme}_run.md`")
    summary = scheme_summary(question, scheme)
    scheme_path = PROJECT_ROOT / "03_methods" / question / f"scheme_{scheme}.md"
    method_text = re.sub(r"\s+", " ", summary["idea"]).strip()
    output_text = re.sub(r"\s+", " ", summary["outputs"]).strip()
    text = text.replace("- 问题目标：", f"- 问题目标：执行 `{question}` 方案 `{scheme}`，严格服务题面中 `{question}` 的建模任务。")
    text = text.replace("- 模型方法：", f"- 模型方法：{method_text} 详见 `{rel(scheme_path)}`。")
    text = text.replace("- 输入数据：", "- 输入数据：`01_problem/data_dictionary.md` 中登记的数据；本题当前数据源为 `01_problem/source/CUMCM2025Problems/C题/附件.xlsx`。")
    text = text.replace("- 输出目标：", f"- 输出目标：{output_text}")
    text = text.replace("- 评价指标：", "- 评价指标：按方案文件要求输出交叉验证误差、模型对比指标、非线性关系改进证据和可供 Codex 重绘的图表数据。")
    text = text.replace("- 关键假设：", "- 关键假设：同一孕妇重复检测需避免数据泄漏；孕周字符串需转换为连续周数；随机过程必须固定 seed。")
    text = text.replace("- 禁止修改的边界：", f"- 禁止修改的边界：不允许更换 `{rel(scheme_path)}` 中的样条/GAM 思路，不允许推进 Q2/Q3/Q4。")
    text = text.replace(
        "- 待填写。",
        f"- `01_problem/problem_statement.md`\n- `01_problem/data_dictionary.md`\n- `{rel(scheme_path)}`",
        1,
    )
    write_text(out_path, text)
    return out_path


def command_prepare_schemes(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    if state["stage"] == "INIT" and not args.force:
        raise SystemExit("problem is not loaded; run import-problem first or use --force")
    assert_question_unlocked(state, question, force=args.force)
    template = read_text(PROJECT_ROOT / "03_methods" / "method_scheme_template.md")
    prompt_path = PROJECT_ROOT / "03_methods" / question / "codex_scheme_generation_prompt.md"
    prompt = render_scheme_generation_prompt(question)
    write_text(prompt_path, prompt)
    generated = [prompt_path]
    for scheme in SCHEMES:
        out_path = PROJECT_ROOT / "03_methods" / question / f"scheme_{scheme}.md"
        if out_path.exists() and not args.overwrite:
            generated.append(out_path)
            state["questions"][question]["schemes"][scheme]["status"] = "draft_ready"
            continue
        text = template
        text = text.replace("- 问题：Q1 / Q2 / Q3", f"- 问题：{question}")
        text = text.replace("- 问题：QX", f"- 问题：{question}")
        text = text.replace("- 方案：A / B / C", f"- 方案：{scheme}")
        text = text.replace("- 方案：X", f"- 方案：{scheme}")
        text = text.replace("- 定位：稳健解释型 / 竞赛均衡型 / 冲奖增强型", f"- 定位：{scheme_position(scheme)}")
        text = text.replace("- 状态：待审批 / 已审批 / 已实现 / 已淘汰", "- 状态：待 Codex 生成")
        write_text(out_path, text)
        generated.append(out_path)
        state["questions"][question]["schemes"][scheme]["status"] = "draft_template"
    state["current_question"] = question
    state["questions"][question]["schemes_generated"] = True
    state["questions"][question]["status"] = "schemes_prepared"
    set_stage(state, "SCHEMES_GENERATED")
    for path in generated:
        append_artifact(state, "scheme_artifact", path, f"{question} scheme workflow artifact")
    save_state(state)
    for path in generated:
        print(path)


def scheme_position(scheme: str) -> str:
    return {"A": "稳健解释型", "B": "竞赛均衡型", "C": "冲奖增强型"}[scheme]


def render_scheme_generation_prompt(question: str) -> str:
    state = load_state()
    prev = previous_question(state, question)
    previous_context = (
        f"""
前置问题依赖：

- `{question}` 必须读取 `{prev}` 的已确认模型、Codex 审查和结果文件。
- 重点读取 `03_methods/{prev}/approved.md`、`06_results/{prev}/` 和相关完成报告。
- 若 `{question}` 已存在历史预分析方案，只能作为草稿参考，必须根据 `{prev}` 的最终结果重审或重写。
"""
        if prev
        else "前置问题依赖：无，本问是当前题目的第一问。"
    )
    return f"""# Codex Scheme Generation Prompt: {question}

你是数学建模多 Agent 系统中的 Codex Agent。请读取：

- `00_shared/WORKFLOW_PROTOCOL.md`
- `00_shared/PROJECT_STATE.md`
- `01_problem/problem_statement.md`
- `01_problem/data_dictionary.md`
- `02_references/paper_style_guide.md`
- `02_references/scoring_rubric.md`
- `03_methods/method_scheme_template.md`

{previous_context}

任务：为 `{question}` 生成 A/B/C 三套可执行方案，并分别写入：

- `03_methods/{question}/scheme_A.md`
- `03_methods/{question}/scheme_B.md`
- `03_methods/{question}/scheme_C.md`

硬性要求：

1. A/B/C 必须有实质差异，不得只是换算法名。
2. 每套方案都要包含数学模型、数据需求、预期图表、敏感性分析、误差分析、实现风险和 Claude Code 实现提示。
3. 推荐 Python，但最终语言以用户审批为准。
4. 不得直接确认最终模型，必须等待用户审批。
5. 如题目信息不足，将边界问题写入 `00_shared/QUESTION_BOUNDARIES.md`。
"""


def command_create_approval_brief(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    assert_question_unlocked(state, question, force=args.force)
    out_path = PROJECT_ROOT / "03_methods" / question / "approval_brief.md"
    text = render_approval_brief(state, question)
    write_text(out_path, text)
    append_artifact(state, "approval_brief", out_path, f"{question} user-facing scheme approval brief")
    save_state(state)
    print_user_facing_brief(out_path, text, f"{question} scheme approval options")


def command_create_model_confirmation_brief(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    fallback_scheme = ensure_scheme(args.scheme) if args.scheme else "B"
    qstate = state["questions"][question]
    passed = [scheme for scheme in SCHEMES if qstate["schemes"][scheme].get("review_result") == "PASS"]
    if not passed and not args.force:
        raise SystemExit(f"{question} has no PASS review result; cannot create model confirmation brief")
    out_path = PROJECT_ROOT / "03_methods" / question / "model_confirmation_brief.md"
    text = render_model_confirmation_brief(state, question, passed[0] if passed else fallback_scheme)
    write_text(out_path, text)
    append_artifact(state, "model_confirmation_brief", out_path, f"{question} model confirmation options")
    save_state(state)
    print_user_facing_brief(out_path, text, f"{question} model confirmation options")


def render_approval_brief(state: dict, question: str) -> str:
    prev = previous_question(state, question)
    dependency_note = (
        f"- 前置依赖：需等待 `{prev}` 模型确认后再生成或修订本问方案。"
        if prev
        else "- 前置依赖：无，本问是当前题目的第一问。"
    )
    rows = []
    details = []
    for scheme in SCHEMES:
        summary = scheme_summary(question, scheme)
        rows.append(
            f"| {scheme} | {summary['position']} | {summary['idea_one_line']} | {summary['pros_one_line']} | {summary['risks_one_line']} |"
        )
        details.extend(
            [
                f"### 方案 {scheme}：{summary['position']}",
                "",
                "**建模思路**",
                "",
                summary["idea"],
                "",
                "**预期输出**",
                "",
                summary["outputs"],
                "",
                "**优点**",
                "",
                summary["pros"],
                "",
                "**风险**",
                "",
                summary["risks"],
                "",
            ]
        )
    return "\n".join(
        [
            f"# {question} 三方案审批简报",
            "",
            "## 当前审批对象",
            "",
            f"- 当前只审批 `{question}`，后续问题暂不进入 Claude Code 执行。",
            dependency_note,
            "- 用户可以选择 A/B/C 中一个方案执行，也可以明确多选；未选方案会保留为备选但不生成 Claude Code 工单。",
            "- 用户确认后，Codex 会生成一段可直接发给 Claude Code 的执行提示词。",
            "",
            "## 三套方案一览",
            "",
            "| 方案 | 定位 | 核心思路 | 主要价值 | 主要风险 |",
            "|---|---|---|---|---|",
            *rows,
            "",
            "## 方案详情",
            "",
            *details,
            "## Codex 初步建议",
            "",
            "- 若用户没有额外偏好，优先考虑 B 方案作为主力竞赛方案。",
            "- A 方案适合作为解释性基准；C 方案适合作为风险、阈值或鲁棒性补强。",
            "- 最终以用户审批为准，Codex 不在此阶段替用户锁定最终模型。",
            "",
            "## 用户回复模板",
            "",
            "```text",
            f"选择 {question} 方案 B。理由：主力方案，先交给 Claude Code 执行。",
            "```",
            "",
            "或：",
            "",
            "```text",
            f"选择 {question} 方案 A,C。理由：先跑解释基准和风险补强。",
            "```",
            "",
        ]
    )


def render_model_confirmation_brief(state: dict, question: str, scheme: str) -> str:
    qstate = state["questions"][question]
    scheme_state = qstate["schemes"][scheme]
    review_path = scheme_state.get("review") or f"06_results/{question}/logs/scheme_{scheme}_codex_review.md"
    completion_path = scheme_state.get("completion_report") or f"04_claude_workorders/completions/{question}_scheme_{scheme}_completion.md"
    metrics = read_metrics_summary(question, scheme)
    metric_lines = [
        f"- Codex 复审结论：`{scheme_state.get('review_result', '待审查')}`",
        f"- 完成报告：`{completion_path}`",
        f"- Codex 审查：`{review_path}`",
    ]
    if metrics:
        metric_lines.extend(
            [
                f"- CV-RMSE：{metrics.get('cv_rmse_spline', 'NA')}",
                f"- CV-R2：{metrics.get('cv_r2_spline', 'NA')}",
                f"- 线性基线 CV-RMSE：{metrics.get('cv_rmse_linear', 'NA')}",
                f"- 线性基线 CV-R2：{metrics.get('cv_r2_linear', 'NA')}",
                f"- 样本量/孕妇数：{metrics.get('n_samples', 'NA')} / {metrics.get('n_subjects', 'NA')}",
            ]
        )
    return "\n".join(
        [
            f"# {question} 模型确认审批简报",
            "",
            "## 当前审批对象",
            "",
            f"- 当前只审批 `{question}` 的最终模型是否确认。",
            f"- 候选模型：方案 `{scheme}`。",
            "- 模型确认后，后续问题才允许正式推进；但图表审批和论文入文仍需单独审批。",
            "",
            "## 复审摘要",
            "",
            *metric_lines,
            "",
            "## 审批选项",
            "",
            "| 选项 | 结论 | 适用情况 | 后续动作 |",
            "|---|---|---|---|",
            f"| 1 | 标准批准 `{question}-{scheme}` | 你接受当前模型和结果，可直接作为本问最终模型 | Codex 记录模型确认，进入中文最终图生成与图表审批 |",
            f"| 2 | 带论文约束批准 `{question}-{scheme}`（推荐） | 你接受模型，但要求论文中保守说明低解释度、残差非正态或高 BMI 不确定性 | Codex 记录模型确认，并把注意事项写入 `approved.md` 和后续论文小节 |",
            "| 3 | 不批准，返修或重跑备选方案 | 你不接受当前模型，认为还需修正、补强或尝试其他方案 | Codex 生成返修要求或重新给出备选方案审批 |",
            "",
            "## Codex 建议",
            "",
            f"- 建议选择 **选项 2**：`{question}-{scheme}` 已通过复审，可以作为本问最终模型；但论文中必须明确模型主要用于趋势刻画，不夸大预测精度。",
            "- 选择选项 2 后，下一步不是直接写论文，而是由 Codex 生成中文最终图，再进入图表审批。",
            "",
            "## 你可以这样回复",
            "",
            "```text",
            f"模型确认选择 {question} 选项 2。理由：接受方案 {scheme} 作为最终模型，但论文中需说明 CV-R2 较低、残差非正态和高 BMI 区域不确定性。",
            "```",
            "",
            "或：",
            "",
            "```text",
            f"模型确认选择 {question} 选项 3。理由：我不满意当前模型，需要 Codex 重新生成返修方案。",
            "```",
            "",
        ]
    )


def read_metrics_summary(question: str, scheme: str) -> dict[str, str]:
    path = PROJECT_ROOT / "06_results" / question / "tables" / f"scheme_{scheme}_metrics.csv"
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return rows[0] if rows else {}


def scheme_summary(question: str, scheme: str) -> dict[str, str]:
    path = PROJECT_ROOT / "03_methods" / question / f"scheme_{scheme}.md"
    text = read_text(path) if path.exists() else ""
    position = find_metadata(text, "定位") or scheme_position(scheme)
    idea = compact_markdown_section(text, "建模思路")
    outputs = compact_markdown_section(text, "预期输出")
    pros = compact_markdown_section(text, "优点")
    risks = compact_markdown_section(text, "风险")
    return {
        "position": position,
        "idea": idea,
        "outputs": outputs,
        "pros": pros,
        "risks": risks,
        "idea_one_line": first_sentence(idea),
        "pros_one_line": first_sentence(pros),
        "risks_one_line": first_sentence(risks),
    }


def find_metadata(text: str, key: str) -> str:
    match = re.search(rf"^- {re.escape(key)}：(.+)$", text, flags=re.MULTILINE)
    return match.group(1).strip() if match else ""


def compact_markdown_section(text: str, title: str, max_lines: int = 8) -> str:
    section = extract_section(text, title)
    if section.startswith(f"## {title}\n未找到") or section == f"## {title}\n未找到。":
        return "待补充。"
    lines = section.splitlines()[1:]
    cleaned = [line.rstrip() for line in lines if line.strip()]
    return "\n".join(cleaned[:max_lines]) if cleaned else "待补充。"


def first_sentence(text: str, max_len: int = 48) -> str:
    plain = re.sub(r"[*`$|#>-]", "", text)
    plain = re.sub(r"\s+", " ", plain).strip()
    for sep in ["。", "；", ";", "."]:
        if sep in plain:
            plain = plain.split(sep)[0]
            break
    if len(plain) > max_len:
        plain = plain[: max_len - 1] + "…"
    return plain or "待补充"


def command_create_claude_prompt(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    assert_question_unlocked(state, question, force=args.force)
    qscheme = state["questions"][question]["schemes"][scheme]
    if qscheme.get("status") not in SELECTED_SCHEME_STATUSES and not args.force:
        raise SystemExit(f"{question} scheme {scheme} is not approved; run approve-schemes first")
    workorder_value = qscheme.get("workorder")
    if not workorder_value:
        workorder_path = create_workorder(question=question, scheme=scheme)
    else:
        workorder_path = PROJECT_ROOT / workorder_value
    prompt_path = PROJECT_ROOT / "04_claude_workorders" / f"{question}_scheme_{scheme}_claude_prompt.md"
    prompt = render_claude_prompt(question, scheme, workorder_path)
    write_text(prompt_path, prompt)
    qscheme["workorder"] = rel(workorder_path)
    qscheme["claude_prompt"] = rel(prompt_path)
    if qscheme.get("status") == "approved_to_run":
        qscheme["status"] = "workorder_created"
    selected = approved_schemes(state["questions"][question]) or [scheme]
    state["questions"][question]["workorders_created"] = all(
        state["questions"][question]["schemes"][s].get("workorder") for s in selected
    )
    state["questions"][question]["status"] = "workorders_created"
    set_stage(state, "CLAUDE_WORKORDERS_CREATED")
    append_artifact(state, "claude_workorder", workorder_path, f"{question} scheme {scheme}")
    append_artifact(state, "claude_prompt", prompt_path, f"{question} scheme {scheme}")
    save_state(state)
    print(prompt_path)


def render_claude_prompt(question: str, scheme: str, workorder_path: Path) -> str:
    scheme_path = PROJECT_ROOT / "03_methods" / question / f"scheme_{scheme}.md"
    completion_path = PROJECT_ROOT / "04_claude_workorders" / "completions" / f"{question}_scheme_{scheme}_completion.md"
    return f"""# Prompt for Claude Code: {question} Scheme {scheme}

你是 Math Magic 多 Agent 数学建模流程中的 Claude Code。请只执行本轮被用户批准的 `{question}` 方案 `{scheme}`，不要推进其他问题。

## 必须先读取

- `00_shared/WORKFLOW_PROTOCOL.md`
- `00_shared/PROJECT_STATE.md`
- `00_shared/QUESTION_BOUNDARIES.md`
- `01_problem/problem_statement.md`
- `01_problem/data_dictionary.md`
- `{rel(scheme_path)}`
- `{rel(workorder_path)}`

## 执行边界

- 你只负责代码实现、运行、调试和结果输出。
- 不允许修改 `{rel(scheme_path)}` 中的建模路线。
- 不允许修改 `03_methods/**/approved.md`、`00_shared/DECISION_LOG.md` 或论文最终结论。
- 如果发现方案不可实现、字段缺失、指标冲突或边界不确定，请写入完成报告的 blocker 区，不要自行换模型。
- 当前只执行 `{question}`，不要生成或运行 Q2/Q3/Q4 的代码。

## 任务

1. 在 `05_code/` 下创建或修改可复现脚本。
2. 从项目根目录运行脚本，固定随机种子。
3. 按工单输出表格、基础图表或绘图数据到 `06_results/{question}/`。
4. 记录完整运行命令、依赖、输入文件、输出文件和关键结果。
5. 完成后写 Markdown 报告到 `{rel(completion_path)}`。

## 完成报告必须包含

- 修改文件清单。
- 运行命令。
- 核心结果表和图的路径。
- 是否完全遵守 `{question}` 方案 `{scheme}`。
- blocker 或需要 Codex/用户决策的问题。
- 可供 Codex 审查的结论摘要。
"""


def command_approve_schemes(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    assert_question_unlocked(state, question, force=args.force)
    schemes = parse_schemes(args.schemes)
    decision_id = args.decision_id or next_decision_id()
    state["questions"][question]["schemes_approved"] = True
    state["questions"][question]["scheme_decision_id"] = decision_id
    state["questions"][question]["status"] = "schemes_approved"
    for scheme in schemes:
        state["questions"][question]["schemes"][scheme]["status"] = "approved_to_run"
        mark_scheme_approval(question, scheme, decision_id, args.notes)
    for scheme in SCHEMES:
        if scheme not in schemes and state["questions"][question]["schemes"][scheme].get("status") in [
            "draft_ready",
            "draft_template",
            "approved_to_run",
        ]:
            state["questions"][question]["schemes"][scheme]["status"] = "not_selected"
    set_stage(state, "SCHEMES_APPROVED")
    save_state(state)
    append_decision(decision_id, f"{question} 方案审批", question, f"批准运行方案：{', '.join(schemes)}", args.notes)
    print(decision_id)


def parse_schemes(value: str) -> list[str]:
    schemes = [ensure_scheme(s.strip()) for s in value.split(",") if s.strip()]
    if not schemes:
        raise SystemExit("at least one scheme is required")
    return schemes


def next_decision_id() -> str:
    log = read_text(PROJECT_ROOT / "00_shared" / "DECISION_LOG.md")
    nums = [int(m.group(1)) for m in re.finditer(r"## D-(\d+)", log)]
    return f"D-{max(nums, default=0) + 1:03d}"


def append_decision(decision_id: str, title: str, scope: str, conclusion: str, notes: str = "") -> None:
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "DECISION_LOG.md",
        f"{decision_id} {title}",
        [
            f"- 日期：{today()}",
            "- 决策人：用户",
            f"- 相关问题：{scope}",
            f"- 结论：{conclusion}",
            f"- 选择理由：{notes or '用户审批通过。'}",
            "- 后续影响：进入下一阶段。",
            "- 是否允许重开：是，需用户重新确认。",
        ],
    )


def mark_scheme_approval(question: str, scheme: str, decision_id: str, notes: str) -> None:
    path = PROJECT_ROOT / "03_methods" / question / f"scheme_{scheme}.md"
    if not path.exists():
        return
    append_markdown_log(
        path,
        "用户审批记录",
        [
            f"- 日期：{today()}",
            f"- 是否批准运行：是",
            f"- 决策 ID：{decision_id}",
            f"- 审批意见：{notes or '批准进入 Claude Code 实现。'}",
        ],
    )


def command_confirm_model(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    if state["questions"][question]["schemes"][scheme].get("review_result") != "PASS" and not args.force:
        raise SystemExit("selected scheme has not passed Codex review; run mark-reviewed --result PASS first or use --force")
    decision_id = args.decision_id or next_decision_id()
    approved_path = PROJECT_ROOT / "03_methods" / question / "approved.md"
    text = [
        f"# {question} Approved Method",
        "",
        "## 审批状态",
        "",
        "- 方案审批：已完成",
        "- 模型确认：已完成",
        f"- 图表审批：{'已完成' if state['questions'][question]['figures_approved'] else '未完成'}",
        "",
        "## 最终采用方案",
        "",
        f"- 方案：{scheme}",
        f"- 决策 ID：{decision_id}",
        f"- 确认日期：{today()}",
        f"- 确认意见：{args.notes or '用户确认该方案作为最终模型。'}",
        "",
        "## 不可修改边界",
        "",
        "- Claude Code 不允许更换模型路线。",
        "- 后续修改必须重新提交用户审批。",
        "",
        "## 论文写入要点",
        "",
        "- Codex 根据最终结果补充。",
        "",
    ]
    write_text(approved_path, "\n".join(text))
    qstate = state["questions"][question]
    qstate["model_confirmed"] = True
    qstate["confirmed_scheme"] = scheme
    qstate["model_decision_id"] = decision_id
    qstate["status"] = "model_confirmed"
    questions = active_question_ids(state)
    if question in questions:
        index = questions.index(question)
        if index + 1 < len(questions):
            next_question = questions[index + 1]
            state["current_question"] = next_question
            if state["questions"][next_question]["status"] == "deferred_waiting_previous_question":
                state["questions"][next_question]["status"] = "active"
    set_stage(state, "MODEL_CONFIRMED")
    append_artifact(state, "approved_method", approved_path, f"{question} confirmed scheme {scheme}")
    save_state(state)
    append_decision(decision_id, f"{question} 最终模型确认", question, f"最终采用方案 {scheme}", args.notes)
    print(approved_path)


def command_create_workorders(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    schemes = parse_schemes(args.schemes)
    assert_question_unlocked(state, question, force=args.force)
    if not state["questions"][question]["schemes_approved"] and not args.force:
        raise SystemExit("schemes are not approved; run approve-schemes first or use --force")
    created = []
    for scheme in schemes:
        if state["questions"][question]["schemes"][scheme].get("status") not in SELECTED_SCHEME_STATUSES and not args.force:
            raise SystemExit(f"{question} scheme {scheme} is not selected/approved")
        out = create_workorder(question=question, scheme=scheme)
        created.append(out)
        state["questions"][question]["schemes"][scheme]["workorder"] = rel(out)
        state["questions"][question]["schemes"][scheme]["status"] = "workorder_created"
        append_artifact(state, "claude_workorder", out, f"{question} scheme {scheme}")
    state["questions"][question]["workorders_created"] = True
    state["questions"][question]["status"] = "workorders_created"
    set_stage(state, "CLAUDE_WORKORDERS_CREATED")
    save_state(state)
    for path in created:
        print(path)


def command_approve_language(args: argparse.Namespace) -> None:
    state = load_state()
    decision_id = args.decision_id or next_decision_id()
    state["language"]["recommended"] = args.language
    state["language"]["approved"] = True
    state["language"]["decision_id"] = decision_id
    save_state(state)
    append_decision(decision_id, "代码语言审批", "全局", f"批准默认实现语言：{args.language}", args.notes)
    print(decision_id)


def command_set_active_question(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    if question not in active_question_ids(state):
        raise SystemExit(f"{question} is not an active question for the imported problem")
    state["current_question"] = question
    questions = active_question_ids(state)
    current_index = questions.index(question)
    for index, qid in enumerate(questions):
        qstate = state["questions"][qid]
        if index < current_index:
            continue
        if qid == question:
            if qstate["schemes_generated"]:
                qstate["status"] = "schemes_prepared"
            elif qstate["status"] in ["not_started", "deferred_waiting_previous_question"]:
                qstate["status"] = "active"
            continue
        if args.defer_later and not qstate.get("model_confirmed"):
            qstate["status"] = "deferred_waiting_previous_question"
            qstate["schemes_approved"] = False
            qstate["scheme_decision_id"] = None
            qstate["workorders_created"] = False
            for scheme in SCHEMES:
                scheme_state = qstate["schemes"][scheme]
                if scheme_state.get("status") in ["draft_ready", "draft_template", "not_started"]:
                    scheme_state["status"] = "deferred_draft"
    save_state(state)
    print(question)


def command_ingest_claude_report(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    src = Path(args.report)
    if not src.is_absolute():
        src = PROJECT_ROOT / src
    if not src.exists():
        raise SystemExit(f"report not found: {src}")
    dest = record_claude_completion(state, question, scheme, src)
    save_state(state)
    print(dest)


def command_check_claude(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    found = check_claude_once(
        state=state,
        question=question,
        scheme=scheme,
        ingest=args.ingest,
        create_review=args.create_review,
        require_standard_outputs=args.require_standard_outputs,
        quiet=False,
    )
    save_state(state)
    if not found and args.fail_missing:
        raise SystemExit(1)


def command_watch_claude(args: argparse.Namespace) -> None:
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    start = time.monotonic()
    print(f"watching Claude completion for {question} scheme {scheme}")
    print(f"interval={args.interval}s timeout={args.timeout}s")
    while True:
        with workflow_lock():
            state = load_state()
            found = check_claude_once(
                state=state,
                question=question,
                scheme=scheme,
                ingest=args.ingest,
                create_review=args.create_review,
                require_standard_outputs=args.require_standard_outputs,
                quiet=args.quiet,
            )
            save_state(state)
        if found:
            print(f"Claude completion detected for {question} scheme {scheme}")
            return
        if args.once:
            print(f"Claude completion not found for {question} scheme {scheme}")
            return
        if args.timeout > 0 and time.monotonic() - start >= args.timeout:
            print(f"timeout waiting for {question} scheme {scheme}")
            raise SystemExit(1)
        time.sleep(args.interval)


def command_dispatch_claude(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    assert_question_unlocked(state, question)
    if state["questions"][question]["schemes"][scheme].get("status") not in SELECTED_SCHEME_STATUSES:
        raise SystemExit(f"{question} scheme {scheme} is not approved or under review; cannot dispatch")
    prompt_path = resolve_dispatch_prompt(state, question, scheme, args.prompt, args.revision)
    mode = args.mode
    command = configured_claude_command(args)
    if mode == "auto":
        mode = "terminal"
    if mode == "terminal":
        script_path = write_terminal_dispatch_script(
            state=state,
            question=question,
            scheme=scheme,
            prompt_path=prompt_path,
            permission_mode=args.terminal_permission_mode,
            target_os=current_target_os(),
        )
        open_terminal_script(script_path, terminal_app=args.terminal_app)
        save_state(state)
        print(script_path)
        print(dispatch_terminal_status_path())
        print("opened visible Claude Code terminal; approve Claude requests in that terminal if prompted")
        if args.watch:
            command_watch_claude(
                argparse.Namespace(
                    question=question,
                    scheme=scheme,
                    interval=args.interval,
                    timeout=args.watch_timeout,
                    once=False,
                    ingest=True,
                    create_review=True,
                    require_standard_outputs=args.require_standard_outputs,
                    quiet=False,
                )
            )
        return
    if mode != "cli":
        raise SystemExit(f"unsupported dispatch mode after cleanup: {mode}")
    if not command:
        raise SystemExit(
            "Claude Code CLI command is not configured. Pass --command, set CLAUDE_CODE_COMMAND, "
            "or create 04_claude_workorders/claude_dispatch_config.json."
        )
    log_path = run_claude_command(state, question, scheme, prompt_path, command, timeout=args.timeout)
    save_state(state)
    print(log_path)
    if args.watch:
        command_watch_claude(
            argparse.Namespace(
                question=question,
                scheme=scheme,
                interval=args.interval,
                timeout=args.watch_timeout,
                once=False,
                ingest=True,
                create_review=True,
                require_standard_outputs=args.require_standard_outputs,
                quiet=False,
            )
        )


def terminal_claude_command(permission_mode: str) -> str:
    return f"claude --permission-mode {shlex.quote(permission_mode)}"


def write_terminal_dispatch_script(
    state: dict,
    question: str,
    scheme: str,
    prompt_path: Path,
    permission_mode: str,
    target_os: str = "auto",
) -> Path:
    target_os = ensure_target_os(target_os)
    if target_os == "windows":
        return write_windows_dispatch_script(state, question, scheme, prompt_path, permission_mode)
    return write_posix_dispatch_script(state, question, scheme, prompt_path, permission_mode)


def write_posix_dispatch_script(
    state: dict,
    question: str,
    scheme: str,
    prompt_path: Path,
    permission_mode: str,
) -> Path:
    dispatch_terminal_dir().mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    script_path = dispatch_terminal_dir() / f"{stamp}_{question}_scheme_{scheme}.sh"
    status_path = dispatch_terminal_status_path()
    expected_report = expected_completion_path(question, scheme)
    standard_outputs = expected_standard_outputs(question, scheme)
    command = terminal_claude_command(permission_mode)
    standard_output_list = [rel(path) for path in standard_outputs]
    status = {
        "status": "terminal_script_created",
        "question": question,
        "scheme": scheme,
        "prompt": rel(prompt_path),
        "script": rel(script_path),
        "permission_mode": permission_mode,
        "created_at": now_iso(),
        "expected_completion_report": rel(expected_report),
        "expected_standard_outputs": standard_output_list,
    }
    script = f"""#!/usr/bin/env bash
set -u
cd {shlex.quote(str(PROJECT_ROOT))}

STATUS_PATH={shlex.quote(rel(status_path))}
PROMPT_PATH={shlex.quote(rel(prompt_path))}
QUESTION={shlex.quote(question)}
SCHEME={shlex.quote(scheme)}

write_status() {{
  local status="$1"
  local rc="${{2:-}}"
  python - "$STATUS_PATH" "$status" "$rc" <<'PY'
import datetime as dt
import json
import sys
from pathlib import Path

path = Path(sys.argv[1])
path.parent.mkdir(parents=True, exist_ok=True)
payload = {{
    "status": sys.argv[2],
    "return_code": sys.argv[3],
    "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
    "question": "{question}",
    "scheme": "{scheme}",
    "prompt": "{rel(prompt_path)}",
    "script": "{rel(script_path)}",
    "permission_mode": "{permission_mode}",
    "expected_completion_report": "{rel(expected_report)}",
    "expected_standard_outputs": {json.dumps(standard_output_list, ensure_ascii=False)},
}}
path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\\n", encoding="utf-8")
PY
}}

write_status running
echo "=== Math Magic Claude Code Terminal Dispatch ==="
echo "Question: $QUESTION"
echo "Scheme: $SCHEME"
echo "Prompt: $PROMPT_PATH"
echo "Expected completion: {rel(expected_report)}"
echo "Permission mode: {permission_mode}"
echo
echo "This terminal is the visible Claude Code interface."
echo "Approve or deny Claude requests here if prompted."
echo "Codex will keep watching the completion files from the other chat."
echo

if [ -f "$HOME/miniconda3/etc/profile.d/conda.sh" ]; then
  source "$HOME/miniconda3/etc/profile.d/conda.sh"
  conda activate base
elif command -v conda >/dev/null 2>&1; then
  eval "$(conda shell.bash hook)"
  conda activate base
fi

{command} "$(cat "$PROMPT_PATH")"
RC=$?
write_status finished "$RC"

echo
echo "=== Claude exited with code $RC ==="
echo "Checking expected completion files..."
python 05_code/tools/agentctl.py check-claude --question "$QUESTION" --scheme "$SCHEME" --ingest --create-review --require-standard-outputs || true
echo
echo "Terminal dispatch finished. You can keep this window open for review."
"""
    write_text(script_path, script)
    script_path.chmod(0o755)
    write_text(status_path, json.dumps(status, ensure_ascii=False, indent=2) + "\n")
    scheme_state = state["questions"][question]["schemes"][scheme]
    scheme_state["dispatch_status"] = "terminal_script_created"
    scheme_state["dispatch_mode"] = "terminal"
    scheme_state["dispatch_prompt"] = rel(prompt_path)
    scheme_state["dispatch_terminal_script"] = rel(script_path)
    scheme_state["dispatch_terminal_status"] = rel(status_path)
    append_artifact(state, "claude_dispatch_terminal_script", script_path, f"{question} scheme {scheme}")
    append_artifact(state, "claude_dispatch_terminal_status", status_path, f"{question} scheme {scheme}")
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} Claude dispatch terminal",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl dispatch-claude",
            f"- 用途：生成可见终端 Claude Code 执行脚本 `{question}` 方案 `{scheme}`。",
            f"- 关联文件：`{rel(prompt_path)}`, `{rel(script_path)}`, `{rel(status_path)}`",
            f"- 权限模式：`{permission_mode}`",
        ],
    )
    return script_path


def write_windows_dispatch_script(
    state: dict,
    question: str,
    scheme: str,
    prompt_path: Path,
    permission_mode: str,
) -> Path:
    dispatch_terminal_dir().mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    script_path = dispatch_terminal_dir() / f"{stamp}_{question}_scheme_{scheme}.ps1"
    status_path = dispatch_terminal_status_path()
    expected_report = expected_completion_path(question, scheme)
    standard_outputs = expected_standard_outputs(question, scheme)
    standard_output_list = [rel(path) for path in standard_outputs]
    output_array = "@(" + ", ".join(ps_single_quote(path) for path in standard_output_list) + ")"
    script = f"""Set-StrictMode -Version Latest
Set-Location {ps_single_quote(PROJECT_ROOT)}

$StatusPath = {ps_single_quote(rel(status_path))}
$PromptPath = {ps_single_quote(rel(prompt_path))}
$Question = {ps_single_quote(question)}
$Scheme = {ps_single_quote(scheme)}
$ExpectedOutputs = {output_array}

function Write-TaskStatus {{
  param([string]$Status, [string]$ReturnCode = "")
  $Payload = [ordered]@{{
    status = $Status
    return_code = $ReturnCode
    updated_at = (Get-Date).ToString("s")
    question = {ps_single_quote(question)}
    scheme = {ps_single_quote(scheme)}
    prompt = {ps_single_quote(rel(prompt_path))}
    script = {ps_single_quote(rel(script_path))}
    permission_mode = {ps_single_quote(permission_mode)}
    expected_completion_report = {ps_single_quote(rel(expected_report))}
    expected_standard_outputs = $ExpectedOutputs
  }}
  $Parent = Split-Path -Parent $StatusPath
  if ($Parent) {{ New-Item -ItemType Directory -Force -Path $Parent | Out-Null }}
  $Payload | ConvertTo-Json -Depth 5 | Set-Content -Path $StatusPath -Encoding UTF8
}}

Write-TaskStatus -Status "running"
Write-Host "=== Math Magic Claude Code VS Code Terminal Dispatch ==="
Write-Host "Question: $Question"
Write-Host "Scheme: $Scheme"
Write-Host "Prompt: $PromptPath"
Write-Host "Expected completion: {rel(expected_report)}"
Write-Host "Permission mode: {permission_mode}"
Write-Host ""
Write-Host "This VS Code terminal is the visible Claude Code interface."
Write-Host "Approve or deny Claude requests here if prompted."
Write-Host "Codex will keep watching the completion files from the other chat."
Write-Host ""

$PromptText = Get-Content -Raw -Encoding UTF8 $PromptPath
& claude --permission-mode {permission_mode} $PromptText
$Rc = if ($LASTEXITCODE -ne $null) {{ $LASTEXITCODE }} else {{ 0 }}
Write-TaskStatus -Status "finished" -ReturnCode "$Rc"

Write-Host ""
Write-Host "=== Claude exited with code $Rc ==="
Write-Host "Checking expected completion files..."
python 05_code/tools/agentctl.py check-claude --question "$Question" --scheme "$Scheme" --ingest --create-review --require-standard-outputs
Write-Host ""
Write-Host "Terminal dispatch finished. You can keep this terminal open for review."
"""
    write_text(script_path, script)
    scheme_state = state["questions"][question]["schemes"][scheme]
    scheme_state["dispatch_status"] = "terminal_script_created"
    scheme_state["dispatch_mode"] = "terminal"
    scheme_state["dispatch_prompt"] = rel(prompt_path)
    scheme_state["dispatch_terminal_script"] = rel(script_path)
    scheme_state["dispatch_terminal_status"] = rel(status_path)
    append_artifact(state, "claude_dispatch_terminal_script", script_path, f"{question} scheme {scheme}")
    append_artifact(state, "claude_dispatch_terminal_status", status_path, f"{question} scheme {scheme}")
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} Claude dispatch terminal",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl dispatch-claude",
            f"- 用途：生成 Windows/PowerShell 可见终端 Claude Code 执行脚本 `{question}` 方案 `{scheme}`。",
            f"- 关联文件：`{rel(prompt_path)}`, `{rel(script_path)}`, `{rel(status_path)}`",
            f"- 权限模式：`{permission_mode}`",
        ],
    )
    return script_path


def open_terminal_script(script_path: Path, terminal_app: str) -> None:
    if script_path.suffix.lower() == ".ps1":
        command = f"powershell -NoProfile -ExecutionPolicy Bypass -File {shlex.quote(str(script_path))}"
    else:
        command = f"bash {shlex.quote(str(script_path))}"
    if sys.platform == "darwin":
        app = terminal_app.strip() or "Terminal"
        if app.lower() == "iterm":
            osa = f'''
tell application "iTerm"
  activate
  create window with default profile
  tell current session of current window
    write text {json.dumps(command)}
  end tell
end tell
'''
        else:
            osa = f'''
tell application "Terminal"
  activate
  do script {json.dumps(command)}
end tell
'''
        result = subprocess.run(["osascript", "-e", osa], cwd=PROJECT_ROOT, check=False)
        if result.returncode != 0:
            raise SystemExit(result.returncode)
        return
    if script_path.suffix.lower() == ".ps1":
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(script_path)],
            cwd=PROJECT_ROOT,
            check=False,
        )
    else:
        result = subprocess.run(["bash", str(script_path)], cwd=PROJECT_ROOT, check=False)
    if result.returncode != 0:
        raise SystemExit(result.returncode)


def command_open_claude_monitor(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    script_path = write_claude_monitor_script(state, question, scheme, interval=args.interval, target_os=args.target_os)
    append_artifact(state, "claude_monitor_terminal_script", script_path, f"{question} scheme {scheme}")
    save_state(state)
    print(script_path)
    if args.no_open:
        return
    open_terminal_script(script_path, terminal_app=args.terminal_app)
    print("opened Claude monitor terminal")


def command_install_vscode_tasks(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    target_os = ensure_target_os(args.target_os)
    if state["questions"][question]["schemes"][scheme].get("status") not in SELECTED_SCHEME_STATUSES:
        raise SystemExit(f"{question} scheme {scheme} is not approved or under review; cannot install dispatch task")
    prompt_path = resolve_dispatch_prompt(state, question, scheme, args.prompt, args.revision)
    dispatch_script = write_terminal_dispatch_script(
        state=state,
        question=question,
        scheme=scheme,
        prompt_path=prompt_path,
        permission_mode=args.terminal_permission_mode,
        target_os=target_os,
    )
    monitor_script = write_claude_monitor_script(state, question, scheme, interval=args.interval, target_os=target_os)
    tasks_path = write_vscode_tasks(question, scheme, dispatch_script, monitor_script, target_os=target_os)
    append_artifact(state, "vscode_tasks", tasks_path, f"{question} scheme {scheme}")
    append_artifact(state, "claude_monitor_terminal_script", monitor_script, f"{question} scheme {scheme}")
    save_state(state)
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} VSCode Claude terminal tasks",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl install-vscode-tasks",
            f"- 用途：为 `{question}` 方案 `{scheme}` 生成 VS Code 集成终端任务。",
            f"- 关联文件：`{rel(tasks_path)}`, `{rel(dispatch_script)}`, `{rel(monitor_script)}`",
            f"- 权限模式：`{args.terminal_permission_mode}`",
            f"- 目标系统：`{target_os}`",
        ],
    )
    print(tasks_path)
    print(f"Run VS Code task: Math Magic: Claude {question}-{scheme} visible session")


def write_vscode_tasks(question: str, scheme: str, dispatch_script: Path, monitor_script: Path, target_os: str) -> Path:
    path = vscode_tasks_path()
    if path.exists():
        try:
            payload = json.loads(read_text(path))
        except json.JSONDecodeError as exc:
            raise SystemExit(f"cannot update invalid VS Code tasks JSON: {path}: {exc}") from exc
    else:
        payload = {"version": "2.0.0", "tasks": []}
    payload["version"] = payload.get("version") or "2.0.0"
    tasks = payload.setdefault("tasks", [])
    dispatch_label = f"Math Magic: Claude {question}-{scheme} dispatch"
    monitor_label = f"Math Magic: Claude {question}-{scheme} monitor"
    watch_label = f"Math Magic: Claude {question}-{scheme} watch"
    session_label = f"Math Magic: Claude {question}-{scheme} visible session"
    managed_labels = {dispatch_label, monitor_label, watch_label, session_label}
    tasks[:] = [task for task in tasks if task.get("label") not in managed_labels]
    tasks.extend(
        [
            vscode_shell_task(
                label=monitor_label,
                command=vscode_script_command(target_os),
                args=vscode_script_args(monitor_script, target_os),
                dedicated_panel=True,
                focus=False,
            ),
            vscode_shell_task(
                label=dispatch_label,
                command=vscode_script_command(target_os),
                args=vscode_script_args(dispatch_script, target_os),
                dedicated_panel=True,
                focus=True,
            ),
            vscode_shell_task(
                label=watch_label,
                command="python",
                args=[
                    "05_code/tools/agentctl.py",
                    "watch-claude",
                    "--question",
                    question,
                    "--scheme",
                    scheme,
                    "--interval",
                    "15",
                    "--ingest",
                    "--create-review",
                    "--require-standard-outputs",
                ],
                dedicated_panel=True,
                focus=False,
            ),
            {
                "label": session_label,
                "dependsOrder": "parallel",
                "dependsOn": [monitor_label, dispatch_label],
                "problemMatcher": [],
                "presentation": {
                    "reveal": "always",
                    "panel": "dedicated",
                    "clear": True,
                    "focus": True,
                },
            },
        ]
    )
    write_text(path, json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    return path


def vscode_script_command(target_os: str) -> str:
    return "powershell" if target_os == "windows" else "bash"


def vscode_script_args(script_path: Path, target_os: str) -> list[str]:
    if target_os == "windows":
        return ["-NoProfile", "-ExecutionPolicy", "Bypass", "-File", rel(script_path)]
    return [rel(script_path)]


def vscode_shell_task(
    label: str,
    command: str,
    args: list[str],
    dedicated_panel: bool,
    focus: bool,
) -> dict:
    return {
        "label": label,
        "type": "shell",
        "command": command,
        "args": args,
        "options": {"cwd": "${workspaceFolder}"},
        "presentation": {
            "reveal": "always",
            "panel": "dedicated" if dedicated_panel else "shared",
            "clear": True,
            "focus": focus,
        },
        "problemMatcher": [],
    }


def write_claude_monitor_script(state: dict, question: str, scheme: str, interval: float, target_os: str = "auto") -> Path:
    target_os = ensure_target_os(target_os)
    if target_os == "windows":
        return write_windows_monitor_script(state, question, scheme, interval)
    return write_posix_monitor_script(state, question, scheme, interval)


def write_posix_monitor_script(state: dict, question: str, scheme: str, interval: float) -> Path:
    dispatch_monitor_dir().mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    script_path = dispatch_monitor_dir() / f"{stamp}_{question}_scheme_{scheme}_monitor.sh"
    qstate = state["questions"][question]
    scheme_state = qstate["schemes"][scheme]
    prompt_path = (
        scheme_state.get("dispatch_prompt")
        or scheme_state.get("claude_prompt")
        or f"04_claude_workorders/{question}_scheme_{scheme}_claude_prompt.md"
    )
    status_path = dispatch_terminal_status_path()
    report_path = expected_completion_path(question, scheme)
    standard_outputs = expected_standard_outputs(question, scheme)
    output_array = "\n".join(f'  "{rel(path)}"' for path in standard_outputs)
    interval_seconds = max(1, int(interval))
    script = f"""#!/usr/bin/env bash
set -u
cd {shlex.quote(str(PROJECT_ROOT))}

QUESTION={shlex.quote(question)}
SCHEME={shlex.quote(scheme)}
PROMPT_PATH={shlex.quote(prompt_path)}
STATUS_PATH={shlex.quote(rel(status_path))}
REPORT_PATH={shlex.quote(rel(report_path))}
INTERVAL={interval_seconds}
STANDARD_OUTPUTS=(
{output_array}
)

while true; do
  clear || true
  echo "=== Math Magic Claude Monitor ==="
  date "+%Y-%m-%d %H:%M:%S"
  echo
  echo "Task: $QUESTION scheme $SCHEME"
  echo "Prompt: $PROMPT_PATH"
  echo
  echo "Claude terminal status:"
  if [ -f "$STATUS_PATH" ]; then
    cat "$STATUS_PATH"
  else
    echo "No terminal status file yet. Dispatch Claude with agentctl dispatch-claude --mode auto."
  fi
  echo
  echo "Completion report:"
  if [ -s "$REPORT_PATH" ]; then
    echo "FOUND: $REPORT_PATH"
    echo
    tail -n 36 "$REPORT_PATH"
  else
    echo "WAITING: $REPORT_PATH"
  fi
  echo
  echo "Standard outputs:"
  for output in "${{STANDARD_OUTPUTS[@]}}"; do
    if [ -s "$output" ]; then
      printf "  [OK] %s\\n" "$output"
    else
      printf "  [--] %s\\n" "$output"
    fi
  done
  echo
  echo "Press q then Enter to close. Refreshing every $INTERVAL seconds."
  if read -r -t "$INTERVAL" answer; then
    if [ "$answer" = "q" ] || [ "$answer" = "Q" ]; then
      exit 0
    fi
  fi
done
"""
    write_text(script_path, script)
    script_path.chmod(0o755)
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} Claude monitor terminal",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl open-claude-monitor",
            f"- 用途：打开 `{question}` 方案 `{scheme}` 的可见 Claude 监控终端。",
            f"- 关联文件：`{rel(script_path)}`, `{rel(status_path)}`, `{rel(report_path)}`",
        ],
    )
    return script_path


def write_windows_monitor_script(state: dict, question: str, scheme: str, interval: float) -> Path:
    dispatch_monitor_dir().mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    script_path = dispatch_monitor_dir() / f"{stamp}_{question}_scheme_{scheme}_monitor.ps1"
    scheme_state = state["questions"][question]["schemes"][scheme]
    prompt_path = (
        scheme_state.get("dispatch_prompt")
        or scheme_state.get("claude_prompt")
        or f"04_claude_workorders/{question}_scheme_{scheme}_claude_prompt.md"
    )
    status_path = dispatch_terminal_status_path()
    report_path = expected_completion_path(question, scheme)
    standard_outputs = expected_standard_outputs(question, scheme)
    output_array = "@(" + ", ".join(ps_single_quote(rel(path)) for path in standard_outputs) + ")"
    interval_seconds = max(1, int(interval))
    script = f"""Set-StrictMode -Version Latest
Set-Location {ps_single_quote(PROJECT_ROOT)}

$Question = {ps_single_quote(question)}
$Scheme = {ps_single_quote(scheme)}
$PromptPath = {ps_single_quote(prompt_path)}
$StatusPath = {ps_single_quote(rel(status_path))}
$ReportPath = {ps_single_quote(rel(report_path))}
$Interval = {interval_seconds}
$StandardOutputs = {output_array}

while ($true) {{
  Clear-Host
  Write-Host "=== Math Magic Claude Monitor ==="
  Write-Host (Get-Date -Format "yyyy-MM-dd HH:mm:ss")
  Write-Host ""
  Write-Host "Task: $Question scheme $Scheme"
  Write-Host "Prompt: $PromptPath"
  Write-Host ""
  Write-Host "Claude terminal status:"
  if (Test-Path $StatusPath) {{
    Get-Content -Raw -Encoding UTF8 $StatusPath
  }} else {{
    Write-Host "No terminal status file yet. Run the VS Code dispatch task first."
  }}
  Write-Host ""
  Write-Host "Completion report:"
  if ((Test-Path $ReportPath) -and ((Get-Item $ReportPath).Length -gt 0)) {{
    Write-Host "FOUND: $ReportPath"
    Write-Host ""
    Get-Content -Encoding UTF8 $ReportPath -Tail 36
  }} else {{
    Write-Host "WAITING: $ReportPath"
  }}
  Write-Host ""
  Write-Host "Standard outputs:"
  foreach ($Output in $StandardOutputs) {{
    if ((Test-Path $Output) -and ((Get-Item $Output).Length -gt 0)) {{
      Write-Host "  [OK] $Output"
    }} else {{
      Write-Host "  [--] $Output"
    }}
  }}
  Write-Host ""
  Write-Host "Press Ctrl+C to close. Refreshing every $Interval seconds."
  Start-Sleep -Seconds $Interval
}}
"""
    write_text(script_path, script)
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} Claude monitor terminal",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl open-claude-monitor",
            f"- 用途：生成 `{question}` 方案 `{scheme}` 的 Windows/PowerShell Claude 监控终端脚本。",
            f"- 关联文件：`{rel(script_path)}`, `{rel(status_path)}`, `{rel(report_path)}`",
        ],
    )
    return script_path


def run_claude_command(
    state: dict,
    question: str,
    scheme: str,
    prompt_path: Path,
    command: str,
    timeout: float,
) -> Path:
    dispatch_log_dir().mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = dispatch_log_dir() / f"{stamp}_{question}_scheme_{scheme}.log"
    args = shlex.split(command)
    if not args:
        raise SystemExit("empty Claude Code command")
    executable = shutil.which(args[0])
    if executable is None:
        raise SystemExit(f"Claude Code command not found: {args[0]}")
    prompt_text = read_text(prompt_path)
    result = subprocess.run(
        args,
        input=prompt_text,
        text=True,
        capture_output=True,
        cwd=PROJECT_ROOT,
        timeout=timeout if timeout > 0 else None,
        check=False,
    )
    log = [
        f"# Claude Dispatch Log: {question} Scheme {scheme}",
        "",
        f"- 时间：{now_iso()}",
        f"- 命令：`{command}`",
        f"- 提示词：`{rel(prompt_path)}`",
        f"- 返回码：{result.returncode}",
        "",
        "## stdout",
        "",
        "```text",
        result.stdout.rstrip(),
        "```",
        "",
        "## stderr",
        "",
        "```",
        result.stderr.rstrip(),
        "```",
        "",
    ]
    write_text(log_path, "\n".join(log))
    scheme_state = state["questions"][question]["schemes"][scheme]
    scheme_state["dispatch_status"] = "sent" if result.returncode == 0 else "failed"
    scheme_state["dispatch_mode"] = "cli"
    scheme_state["dispatch_prompt"] = rel(prompt_path)
    scheme_state["dispatch_log"] = rel(log_path)
    append_artifact(state, "claude_dispatch_log", log_path, f"{question} scheme {scheme}")
    append_markdown_log(
        PROJECT_ROOT / "00_shared" / "AI_USAGE_LOG.md",
        f"AILOG-AUTO-{now_iso()} Claude dispatch cli",
        [
            f"- 日期：{today()}",
            "- 工具：agentctl dispatch-claude",
            f"- 用途：通过 CLI 调用 Claude Code 执行 `{question}` 方案 `{scheme}`。",
            f"- 关联文件：`{rel(prompt_path)}`, `{rel(log_path)}`",
            f"- 返回码：{result.returncode}",
        ],
    )
    if result.returncode != 0:
        raise SystemExit(result.returncode)
    return log_path


def check_claude_once(
    state: dict,
    question: str,
    scheme: str,
    ingest: bool,
    create_review: bool,
    require_standard_outputs: bool,
    quiet: bool,
) -> bool:
    report_path = expected_completion_path(question, scheme)
    standard_outputs = expected_standard_outputs(question, scheme)
    report_found = report_path.exists() and report_path.stat().st_size > 0
    output_status = [(path, path.exists() and path.stat().st_size > 0) for path in standard_outputs]
    outputs_ok = all(ok for _, ok in output_status)
    found = report_found and (outputs_ok if require_standard_outputs else True)
    if not quiet:
        print("== Claude Completion Check ==")
        print(f"question: {question}")
        print(f"scheme: {scheme}")
        print(f"completion_report: {'found' if report_found else 'missing'} {rel(report_path)}")
        for path, ok in output_status:
            print(f"standard_output: {'found' if ok else 'missing'} {rel(path)}")
    if found and ingest:
        dest = record_claude_completion(state, question, scheme, report_path)
        if not quiet:
            print(f"ingested: {rel(dest)}")
    if found and create_review:
        review = create_review_file(state, question, scheme)
        if not quiet:
            print(f"review_template: {rel(review)}")
    return found


def command_create_review(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    out_path = create_review_file(state, question, scheme)
    save_state(state)
    print(out_path)


def render_codex_review_template(question: str, scheme: str, report: str) -> str:
    return f"""# Codex Review: {question} Scheme {scheme}

## 基本信息

- 问题：{question}
- 方案：{scheme}
- Claude 完成报告：`{report}`
- 审查日期：{today()}
- 审查结论：PASS / REVISE / BLOCKED

## 路线一致性

- 是否严格遵守 `03_methods/{question}/scheme_{scheme}.md`：
- 是否偏离已审批边界：
- 是否需要用户仲裁：

## 代码审查

- 可复现性：
- 随机种子：
- 路径处理：
- 数据清洗：
- 异常处理：

## 结果审查

- 输出表格：
- 输出图表：
- 核心指标：
- 是否支持题目结论：

## 数模论文补强

- 敏感性分析：
- 误差分析：
- 鲁棒性检验：
- 对比基准：

## 退回 Claude Code 的修改项

1. 待填写。

## 可进入论文的材料

- 表格：
- 图：
- 文字结论：
"""


def command_mark_reviewed(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    scheme = ensure_scheme(args.scheme)
    status = args.result.upper()
    if status not in ["PASS", "REVISE", "BLOCKED"]:
        raise SystemExit("result must be PASS, REVISE, or BLOCKED")
    state["questions"][question]["schemes"][scheme]["review_result"] = status
    state["questions"][question]["schemes"][scheme]["status"] = f"review_{status.lower()}"
    state["questions"][question]["status"] = f"review_{status.lower()}"
    selected = approved_schemes(state["questions"][question]) or [scheme]
    state["questions"][question]["code_reviewed"] = all(
        state["questions"][question]["schemes"][s].get("review_result") == "PASS" for s in selected
    )
    if state["questions"][question]["code_reviewed"]:
        set_stage(state, "CODE_REVIEWED")
    save_state(state)
    print(f"{question} {scheme}: {status}")


def command_compare_schemes(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    template = read_text(PROJECT_ROOT / "06_results" / "experiment_comparison_template.md")
    out_path = PROJECT_ROOT / "06_results" / question / f"{question}_scheme_comparison.md"
    text = template.replace("- 问题：Q1 / Q2 / Q3", f"- 问题：{question}")
    text = text.replace("- 问题：QX", f"- 问题：{question}")
    write_text(out_path, text)
    append_artifact(state, "scheme_comparison", out_path, question)
    save_state(state)
    print(out_path)


def command_approve_figures(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    decision_id = args.decision_id or next_decision_id()
    qstate = state["questions"][question]
    if args.allow_non_chinese:
        qstate["figure_language"] = "non_chinese_exception"
    else:
        qstate["figure_language"] = "zh"
    qstate["figures_generated"] = True
    qstate["figures_approved"] = True
    qstate["figure_decision_id"] = decision_id
    qstate["status"] = "figures_approved"
    set_stage(state, "FIGURES_APPROVED")
    save_state(state)
    language_note = "非中文例外审批" if args.allow_non_chinese else "中文最终图"
    append_decision(
        decision_id,
        f"{question} 图表审批",
        question,
        f"批准图表（{language_note}）：{args.figures or '见 07_paper/figures 与 06_results'}",
        args.notes,
    )
    print(decision_id)


def command_mark_paper_written(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    state["questions"][question]["paper_written"] = True
    state["questions"][question]["paper_section"] = rel(question_section_path(question))
    active_questions = state["problem"].get("question_ids") or QUESTIONS
    if all(state["questions"][q]["paper_written"] for q in active_questions):
        set_stage(state, "PAPER_WRITTEN")
    save_state(state)
    print(f"{question}: paper_written")


def confirmed_scheme(state: dict, question: str) -> str:
    scheme = state["questions"][question].get("confirmed_scheme")
    if scheme:
        return scheme
    for candidate in SCHEMES:
        if state["questions"][question]["schemes"][candidate].get("review_result") == "PASS":
            return candidate
    return "待确认"


def read_csv_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def fmt_num(value: str | float | int | None, digits: int = 4) -> str:
    if value in (None, ""):
        return "待补充"
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return str(value)


def fmt_sci(value: str | float | int | None) -> str:
    if value in (None, ""):
        return "待补充"
    try:
        return f"{float(value):.2e}"
    except (TypeError, ValueError):
        return str(value)


def render_q1_scheme_b_section() -> str | None:
    table_dir = PROJECT_ROOT / "06_results" / "Q1" / "tables"
    metrics_rows = read_csv_rows(table_dir / "scheme_B_metrics.csv")
    cutoff_rows = read_csv_rows(table_dir / "scheme_B_sensitivity_gestation_cutoff.csv")
    bmi_rows = read_csv_rows(table_dir / "scheme_B_partial_bmi.csv")
    if not metrics_rows:
        return None
    metrics = metrics_rows[0]

    cutoff_25 = next((r for r in cutoff_rows if r.get("dataset") == "<=25w"), {})
    cutoff_26 = next((r for r in cutoff_rows if r.get("dataset") == "<=26w"), {})
    raw_ci_min = ""
    if bmi_rows:
        raw_values = []
        for row in bmi_rows:
            try:
                raw_values.append(float(row.get("ci_lo_raw", "")))
            except ValueError:
                pass
        if raw_values:
            raw_ci_min = fmt_num(min(raw_values), 4)

    replacements = {
        "__N_SAMPLES__": str(int(float(metrics.get("n_samples", 0) or 0))),
        "__N_SUBJECTS__": str(int(float(metrics.get("n_subjects", 0) or 0))),
        "__GEST_RANGE__": metrics.get("gestation_range", "待补充"),
        "__BMI_RANGE__": metrics.get("bmi_range", "待补充"),
        "__BEST_DEGREE__": str(int(float(metrics.get("best_degree", 0) or 0))),
        "__N_KNOTS__": str(int(float(metrics.get("n_knots", 0) or 0))),
        "__ALPHA__": fmt_num(metrics.get("alpha"), 4),
        "__TRAIN_RMSE__": fmt_num(metrics.get("train_rmse"), 4),
        "__TRAIN_R2__": fmt_num(metrics.get("train_r2"), 4),
        "__CV_RMSE__": fmt_num(metrics.get("cv_rmse_spline"), 4),
        "__CV_R2__": fmt_num(metrics.get("cv_r2_spline"), 4),
        "__LIN_RMSE__": fmt_num(metrics.get("cv_rmse_linear"), 4),
        "__LIN_R2__": fmt_num(metrics.get("cv_r2_linear"), 4),
        "__DELTA_RMSE__": fmt_num(metrics.get("delta_rmse"), 4),
        "__SHAPIRO_P__": fmt_sci(metrics.get("residual_shapiro_p")),
        "__PEARSON_R__": fmt_num(metrics.get("residual_pearson_r"), 3),
        "__PEARSON_P__": fmt_num(metrics.get("residual_pearson_p"), 3),
        "__CUTOFF25_RMSE__": fmt_num(cutoff_25.get("cv_rmse")),
        "__CUTOFF25_R2__": fmt_num(cutoff_25.get("cv_r2")),
        "__CUTOFF26_RMSE__": fmt_num(cutoff_26.get("cv_rmse")),
        "__CUTOFF26_R2__": fmt_num(cutoff_26.get("cv_r2")),
        "__RAW_CI_MIN__": raw_ci_min or "待补充",
    }

    template = r"""\subsection{问题一模型的建立与求解}\label{sub:5.1}

问题一要求分析男胎胎儿的 Y 染色体浓度与孕周、孕妇 BMI 等因素之间的关系。由于同一孕妇存在多次检测记录，若直接随机划分训练集和测试集，容易把同一孕妇的个体差异泄漏到验证集中。因此本文在问题一中采用按孕妇代码分组的交叉验证，并在非线性关系刻画上使用样条岭回归模型。

\subsubsection{数据预处理与变量定义}

记第 $i$ 名孕妇第 $j$ 次检测的孕周为 $t_{ij}$，孕妇 BMI 为 $b_i$，男胎 Y 染色体浓度为 $y_{ij}$，测序质量控制变量向量为 $\boldsymbol q_{ij}$。根据题目中男胎检测数据的可解释范围，本文保留孕周在 $10$ 至 $25$ 周内、Y 染色体浓度为正且核心字段完整的样本。清洗后样本量为 __N_SAMPLES__ 条，涉及 __N_SUBJECTS__ 名孕妇；孕周范围为 __GEST_RANGE__ 周，BMI 范围为 __BMI_RANGE__。

\subsubsection{样条岭回归模型}

考虑孕周和 BMI 与 Y 染色体浓度之间可能存在非线性关系，本文采用 B 样条基函数分别展开孕周和 BMI，并加入二者的交互项。模型形式为
\[
  y_{ij}=\beta_0+f_1(t_{ij})+f_2(b_i)+f_3(t_{ij},b_i)+\boldsymbol{\gamma}^{\mathrm T}\boldsymbol q_{ij}+\varepsilon_{ij},
\]
其中 $f_1(\cdot)$ 和 $f_2(\cdot)$ 分别表示孕周与 BMI 的非线性边际效应，$f_3(\cdot,\cdot)$ 表示二者的交互影响，$\varepsilon_{ij}$ 为随机误差项。为降低样条展开后多重共线性对估计稳定性的影响，模型采用岭回归惩罚：
\[
  \min_{\boldsymbol\theta}\sum_{i,j}\left(y_{ij}-\hat y_{ij}\right)^2+\alpha\|\boldsymbol\theta\|_2^2 .
\]

样条阶数在 $3$ 至 $6$ 之间比较，外层采用 5 折 GroupKFold，内层按孕妇分组选择惩罚参数 $\alpha$。最终选择样条阶数 __BEST_DEGREE__、节点数 __N_KNOTS__、惩罚参数 $\alpha=__ALPHA__$ 的模型。建模流程如 \ref{fig:q1_flow} 所示。

\begin{figure}[htbp]
  \centering
  \includegraphics[width=0.76\textwidth]{figures/q1_fig5_model_flow.png}
  \caption{问题一样条岭回归建模流程}
  \label{fig:q1_flow}
\end{figure}

\subsubsection{模型结果与关系分析}

表 \ref{tab:q1_metrics} 给出了样条岭回归模型与线性基线模型的主要评价结果。样条岭回归的交叉验证 RMSE 为 __CV_RMSE__，略低于线性基线的 __LIN_RMSE__；交叉验证 $R^2$ 为 __CV_R2__，高于线性基线的 __LIN_R2__。这说明非线性模型相较线性基线具有一定改进，但解释度仍然有限，模型更适合用于刻画总体趋势和为后续检测时点优化提供关系基础。

\begin{table}[!ht]
  \centering
  \caption{问题一模型主要评价指标}
  \label{tab:q1_metrics}
  \begin{tabularx}{\textwidth}{cYYYY}
    \toprule
    模型 & 训练 RMSE & 训练 $R^2$ & CV-RMSE & CV-$R^2$ \\
    \midrule
    样条岭回归 & __TRAIN_RMSE__ & __TRAIN_R2__ & __CV_RMSE__ & __CV_R2__ \\
    线性岭回归基线 & - & - & __LIN_RMSE__ & __LIN_R2__ \\
    \bottomrule
  \end{tabularx}
\end{table}

\begin{figure}[!ht]
  \centering
  \begin{subfigure}{0.42\textwidth}
    \centering
    \includegraphics[width=\textwidth]{figures/q1_fig1_gestation_effect.png}
    \caption{孕周边际效应}
  \end{subfigure}
  \hspace{0.05\textwidth}
  \begin{subfigure}{0.42\textwidth}
    \centering
    \includegraphics[width=\textwidth]{figures/q1_fig2_bmi_effect.png}
    \caption{BMI 边际效应}
  \end{subfigure}
  \caption{孕周和 BMI 对男胎 Y 染色体浓度的非线性影响}
  \label{fig:q1_effects}
\end{figure}

\FloatBarrier

从 \ref{fig:q1_effects} 可以看出，在控制 BMI 和测序质控变量后，预测的 Y 染色体浓度随孕周变化呈非线性形态：早中孕阶段有上升趋势，中间阶段略有回落，后段再次升高。BMI 的边际效应同样呈非线性，且在 BMI 边界区间置信带明显变宽，说明边界 BMI 样本支持不足，相关结论应谨慎外推。

进一步地，\ref{fig:q1_surface} 展示了孕周与 BMI 联合作用下的预测浓度面。红色曲线对应 $4\%$ 的临床参考浓度线。可以看到，在较低孕周与部分高 BMI 区域中，预测浓度接近或低于参考线；随着孕周增加，多数 BMI 区间的预测浓度上升。这一结果为第二问中根据 BMI 分组确定 NIPT 检测时点提供了直接依据。

\begin{figure}[!ht]
  \centering
  \includegraphics[width=0.56\textwidth]{figures/q1_fig3_gestation_bmi_surface.png}
  \caption{孕周和 BMI 联合作用下的 Y 染色体浓度预测面}
  \label{fig:q1_surface}
\end{figure}

\FloatBarrier

\subsubsection{敏感性与误差分析}

为检验模型稳定性，本文比较了不同样条阶数和不同孕周上界下的结果。样条阶数从 $3$ 增至 $6$ 时，CV-RMSE 的变化幅度很小，最优值出现在三阶样条；将主分析孕周上界从 $25$ 周放宽至 $26$ 周后，CV-RMSE 由 __CUTOFF25_RMSE__ 变为 __CUTOFF26_RMSE__，CV-$R^2$ 由 __CUTOFF25_R2__ 变为 __CUTOFF26_R2__，说明结论对孕周上界设置不敏感。

\begin{table}[htbp]
  \centering
  \caption{问题一模型检验与敏感性摘要}
  \label{tab:q1_validation}
  \begin{tabularx}{\textwidth}{cY}
    \toprule
    检验项 & 结果与解释 \\
    \midrule
    样条阶数敏感性 & 3 至 6 阶 CV-RMSE 变化幅度较小，最优阶数为 __BEST_DEGREE__。 \\
    孕周上界敏感性 & 上界由 25 周放宽至 26 周后，CV-RMSE 为 __CUTOFF26_RMSE__，结论基本稳定。 \\
    残差正态性 & Shapiro 检验 $p=__SHAPIRO_P__$，残差不满足正态假设。 \\
    残差相关性 & 残差与拟合值 Pearson $r=__PEARSON_R__$，$p=__PEARSON_P__$，未见明显线性相关。 \\
    边界不确定性 & BMI 高端样本较少，Bootstrap 原始置信区间下界曾低至 __RAW_CI_MIN__，已按浓度非负约束截断。 \\
    \bottomrule
  \end{tabularx}
\end{table}

误差分析方面，残差 Shapiro 检验的 $p$ 值为 __SHAPIRO_P__，显示残差不满足正态性假设；残差与拟合值的 Pearson 相关系数为 __PEARSON_R__，对应 $p$ 值为 __PEARSON_P__，未显示明显线性相关。BMI 边际效应的 Bootstrap 原始置信区间下界最小约为 __RAW_CI_MIN__，按浓度非负约束截断为 $0$，反映出边界 BMI 区域预测不确定性较高。因此，后续仅将其作为趋势刻画和时点优化的关系模型，不解释为高精度医学单点预测。

\begin{figure}[htbp]
  \centering
  \includegraphics[width=0.66\textwidth]{figures/q1_fig4_spline_sensitivity.png}
  \caption{样条阶数对问题一交叉验证误差的影响}
  \label{fig:q1_degree_sensitivity}
\end{figure}

\FloatBarrier

\subsubsection{本问小结}

综上，问题一建立了包含孕周、BMI、二者交互项和测序质控变量的样条岭回归模型。结果表明，男胎 Y 染色体浓度与孕周、BMI 均存在非线性关系；非线性模型较线性基线有小幅稳定改进，但受个体差异和检测误差影响，整体解释度有限。后续问题将基于该模型给出的孕周-BMI 响应关系，进一步讨论 BMI 分组和最佳 NIPT 检测时点。
"""
    for key, value in replacements.items():
        template = template.replace(key, value)
    return template


def render_question_paper_section(state: dict, question: str) -> str:
    scheme = confirmed_scheme(state, question)
    if question == "Q1" and scheme == "B":
        q1_text = render_q1_scheme_b_section()
        if q1_text:
            return q1_text
    title_num = question[1:] if question[1:].isdigit() else question
    return f"""\\subsection{{问题{title_num}模型的建立与求解}}\\label{{sub:5.{title_num}}}

本节内容由 Codex 在该问题完成模型确认、图表审批后写入。当前采用方案为 {scheme}。正文写入时必须包含变量定义、模型表达式、求解流程、结果表、中文图表及对题目要求的直接回答。

\\subsubsection{{模型建立}}

待 Codex 根据已确认模型补充数学表达式、参数说明和约束条件。

\\subsubsection{{模型求解与结果分析}}

待 Codex 根据 `06_results/{question}/tables/` 中的结果表和已审批中文图表补充。

\\subsubsection{{本问小结}}

待 Codex 概括本问结论，并说明对后续问题的输入或约束影响。
"""


def run_latex_compile() -> Path:
    paper_dir = PROJECT_ROOT / "07_paper"
    result = subprocess.run(
        ["xelatex", "-interaction=nonstopmode", "-halt-on-error", "main.tex"],
        cwd=paper_dir,
        text=True,
        capture_output=True,
        check=False,
    )
    print(result.stdout[-4000:])
    if result.returncode != 0:
        print(result.stderr)
        raise SystemExit(result.returncode)
    return paper_dir / "main.pdf"


def parse_aux_label_pages(aux_path: Path) -> dict[str, int]:
    if not aux_path.exists():
        return {}
    labels: dict[str, int] = {}
    pattern = re.compile(r"\\newlabel\{([^}]+)\}\{\{[^{}]*\}\{(\d+)\}")
    for line in read_text(aux_path).splitlines():
        match = pattern.search(line)
        if match:
            labels[match.group(1)] = int(match.group(2))
    return labels


def parse_label_expectations(expectations: list[str]) -> dict[str, int]:
    parsed: dict[str, int] = {}
    for value in expectations:
        if "=" not in value:
            raise SystemExit(f"invalid label expectation `{value}`; expected label=page")
        label, page = value.split("=", 1)
        label = label.strip()
        try:
            parsed[label] = int(page.strip())
        except ValueError as exc:
            raise SystemExit(f"invalid page in label expectation `{value}`") from exc
    return parsed


def detect_large_pdf_gaps(pdf_path: Path, max_internal_gap_ratio: float) -> list[dict]:
    try:
        import fitz
    except ImportError as exc:
        raise SystemExit("layout-check requires pymupdf; install project dependencies first") from exc

    if not pdf_path.exists():
        raise SystemExit(f"PDF not found: {rel(pdf_path)}")

    document = fitz.open(pdf_path)
    findings: list[dict] = []
    for page_index, page in enumerate(document, start=1):
        pix = page.get_pixmap(matrix=fitz.Matrix(0.55, 0.55), colorspace=fitz.csGRAY, alpha=False)
        width, height = pix.width, pix.height
        samples = pix.samples
        min_ink_pixels = max(4, int(width * 0.012))
        content_rows: list[int] = []
        for row in range(height):
            start = row * width
            end = start + width
            ink = sum(1 for value in samples[start:end] if value < 245)
            if ink >= min_ink_pixels:
                content_rows.append(row)
        if len(content_rows) < 2:
            continue
        top_guard = int(height * 0.08)
        bottom_guard = int(height * 0.92)
        largest_gap = 0
        largest_pair: tuple[int, int] | None = None
        for previous, current in zip(content_rows, content_rows[1:]):
            gap = current - previous - 1
            if gap <= largest_gap:
                continue
            if previous < top_guard or current > bottom_guard:
                continue
            largest_gap = gap
            largest_pair = (previous, current)
        ratio = largest_gap / height if height else 0.0
        if largest_pair and ratio > max_internal_gap_ratio:
            findings.append(
                {
                    "page": page_index,
                    "gap_ratio": ratio,
                    "gap_start_ratio": largest_pair[0] / height,
                    "gap_end_ratio": largest_pair[1] / height,
                }
            )
    return findings


def run_layout_check(
    pdf_path: Path,
    aux_path: Path,
    max_internal_gap_ratio: float,
    label_expectations: list[str],
) -> list[str]:
    issues: list[str] = []
    for finding in detect_large_pdf_gaps(pdf_path, max_internal_gap_ratio):
        issues.append(
            "page {page}: internal blank gap {gap:.1%} from {start:.1%} to {end:.1%}".format(
                page=finding["page"],
                gap=finding["gap_ratio"],
                start=finding["gap_start_ratio"],
                end=finding["gap_end_ratio"],
            )
        )

    expected_pages = parse_label_expectations(label_expectations)
    if expected_pages:
        label_pages = parse_aux_label_pages(aux_path)
        for label, expected_page in expected_pages.items():
            actual_page = label_pages.get(label)
            if actual_page is None:
                issues.append(f"label `{label}` not found in {rel(aux_path)}")
            elif actual_page != expected_page:
                issues.append(f"label `{label}` is on page {actual_page}, expected page {expected_page}")
    return issues


def command_layout_check(args: argparse.Namespace) -> None:
    pdf_path = args.pdf if args.pdf.is_absolute() else PROJECT_ROOT / args.pdf
    aux_path = args.aux if args.aux.is_absolute() else PROJECT_ROOT / args.aux
    issues = run_layout_check(
        pdf_path=pdf_path,
        aux_path=aux_path,
        max_internal_gap_ratio=args.max_internal_gap_ratio,
        label_expectations=args.expect_label_page,
    )
    print("== Paper Layout Check ==")
    print(f"pdf: {rel(pdf_path)}")
    print(f"max_internal_gap_ratio: {args.max_internal_gap_ratio:.2f}")
    if args.expect_label_page:
        print(f"label_expectations: {', '.join(args.expect_label_page)}")
    if issues:
        for issue in issues:
            print(f"FAIL: {issue}")
        raise SystemExit(1)
    print("layout: ok")


def command_write_question_paper(args: argparse.Namespace) -> None:
    state = load_state()
    question = ensure_question(args.question)
    qstate = state["questions"][question]
    if not qstate.get("model_confirmed") and not args.force:
        raise SystemExit(f"{question} model is not confirmed; run confirm-model first or use --force")
    if not qstate.get("figures_approved") and not args.force:
        raise SystemExit(f"{question} figures are not approved; run approve-figures first or use --force")
    section_path = question_section_path(question)
    if section_path.exists() and not args.force:
        current = read_text(section_path)
        if "待用户完成方案审批和模型确认后写入" not in current and "待 Codex 根据已确认模型补充" not in current:
            raise SystemExit(f"{rel(section_path)} already has non-template content; use --force to overwrite")
    write_text(section_path, render_question_paper_section(state, question))
    qstate["paper_written"] = True
    qstate["paper_section"] = rel(section_path)
    qstate["status"] = "paper_written"
    append_artifact(state, "paper_question_section", section_path, question)
    pdf_path = None
    if args.compile:
        pdf_path = run_latex_compile()
        layout_issues = run_layout_check(
            pdf_path=pdf_path,
            aux_path=PROJECT_ROOT / "07_paper" / "main.aux",
            max_internal_gap_ratio=0.24,
            label_expectations=[],
        )
        if layout_issues:
            print("== Paper Layout Check ==")
            for issue in layout_issues:
                print(f"FAIL: {issue}")
            raise SystemExit(1)
        qstate["latex_compiled_at"] = now_iso()
        append_artifact(state, "latex_pdf", pdf_path, f"after {question} paper write")
    active_questions = state["problem"].get("question_ids") or QUESTIONS
    if all(state["questions"][q]["paper_written"] for q in active_questions):
        set_stage(state, "PAPER_WRITTEN")
    save_state(state)
    print(section_path)
    if pdf_path:
        print(pdf_path)


def command_finalize_summary_paper(args: argparse.Namespace) -> None:
    state = load_state()
    active_questions = state["problem"].get("question_ids") or QUESTIONS
    missing = [q for q in active_questions if not state["questions"][q].get("paper_written")]
    if missing and not args.force:
        raise SystemExit(f"summary sections are locked until all question sections are written; missing: {', '.join(missing)}")
    for section in paper_summary_sections():
        if not section.exists():
            raise SystemExit(f"missing summary section: {rel(section)}")
    if args.compile:
        pdf_path = run_latex_compile()
        layout_issues = run_layout_check(
            pdf_path=pdf_path,
            aux_path=PROJECT_ROOT / "07_paper" / "main.aux",
            max_internal_gap_ratio=0.24,
            label_expectations=[],
        )
        if layout_issues:
            print("== Paper Layout Check ==")
            for issue in layout_issues:
                print(f"FAIL: {issue}")
            raise SystemExit(1)
        append_artifact(state, "latex_pdf", pdf_path, "after summary finalize check")
    save_state(state)
    print("summary sections are ready for final Codex writing")


def command_paper_check(args: argparse.Namespace) -> None:
    checks = []
    for path in [
        PROJECT_ROOT / "07_paper" / "main.tex",
        PROJECT_ROOT / "07_paper" / "sections" / "abstract.tex",
        PROJECT_ROOT / "07_paper" / "sections" / "problem_analysis.tex",
        PROJECT_ROOT / "07_paper" / "sections" / "model_validation.tex",
        PROJECT_ROOT / "07_paper" / "appendix" / "code_appendix.tex",
        PROJECT_ROOT / "02_references" / "paper_style_guide.md",
        PROJECT_ROOT / "02_references" / "scoring_rubric.md",
    ]:
        checks.append((rel(path), path.exists()))
    print("== Paper Check ==")
    failed = False
    for path, ok in checks:
        print(f"{path}: {'ok' if ok else 'missing'}")
        failed = failed or not ok
    if failed:
        raise SystemExit(1)


def command_readiness(args: argparse.Namespace) -> None:
    state = load_state()
    active_questions = state["problem"].get("question_ids") or QUESTIONS
    next_question = next_question_to_solve(state)
    print("== Workflow Readiness ==")
    print(f"stage: {state['stage']}")
    print(f"workflow mode: {state.get('workflow', {}).get('mode', 'sequential')}")
    print(f"problem loaded: {'ok' if state['problem']['title'] else 'pending'}")
    print(f"language approved: {'ok' if state['language']['approved'] else 'pending'}")
    print(f"question order: {', '.join(active_questions)}")
    print(f"current unlocked question: {next_question or 'all model_confirmed'}")
    for question in active_questions:
        qstate = state["questions"][question]
        print(
            f"{question}: status={qstate['status']} schemes={qstate['schemes_generated']} approved={qstate['schemes_approved']} "
            f"workorders={qstate['workorders_created']} model={qstate['model_confirmed']} figures={qstate['figures_approved']} paper={qstate['paper_written']}"
        )


def command_latex_check(args: argparse.Namespace) -> None:
    pdf_path = run_latex_compile()
    print(pdf_path)
    if args.skip_layout_check:
        return
    issues = run_layout_check(
        pdf_path=pdf_path,
        aux_path=PROJECT_ROOT / "07_paper" / "main.aux",
        max_internal_gap_ratio=args.max_internal_gap_ratio,
        label_expectations=args.expect_label_page,
    )
    print("== Paper Layout Check ==")
    if issues:
        for issue in issues:
            print(f"FAIL: {issue}")
        raise SystemExit(1)
    print("layout: ok")



def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Math Magic workflow helper")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_state = subparsers.add_parser("init-state", help="create or reset machine-readable workflow state")
    init_state.add_argument("--force", action="store_true")
    init_state.set_defaults(func=command_init_state)

    status = subparsers.add_parser("status", help="print shared project state")
    status.set_defaults(func=command_status)

    env_check = subparsers.add_parser("env-check", help="check local runtime dependencies")
    env_check.set_defaults(func=command_env_check)

    doctor = subparsers.add_parser("doctor", help="run first-install checks and optionally install a VS Code Claude smoke task")
    doctor.add_argument("--target-os", default="auto", choices=["auto", "windows", "posix"])
    doctor.add_argument("--write-vscode-smoke-task", action="store_true")
    doctor.add_argument("--strict", action="store_true")
    doctor.set_defaults(func=command_doctor)

    tools = subparsers.add_parser("tools", help="print registered workflow tools")
    tools.set_defaults(func=command_tools)

    import_problem = subparsers.add_parser("import-problem", help="import a training problem statement")
    import_problem.add_argument("--statement", type=Path, default=None)
    import_problem.add_argument("--title", default=None)
    import_problem.add_argument("--problem-id", default=None)
    import_problem.add_argument("--data-dir", type=Path, default=None)
    import_problem.add_argument("--num-questions", type=int, default=3)
    import_problem.set_defaults(func=command_import_problem)

    scan_data = subparsers.add_parser("scan-data", help="scan raw data files into data_dictionary.md")
    scan_data.add_argument("--data-dir", type=Path, required=True)
    scan_data.add_argument("--out", type=Path, default=None)
    scan_data.add_argument("--sample-rows", type=int, default=200)
    scan_data.add_argument("--include-unsupported", action="store_true")
    scan_data.set_defaults(func=command_scan_data)

    approve_language = subparsers.add_parser("approve-language", help="approve the default implementation language")
    approve_language.add_argument("--language", default="Python")
    approve_language.add_argument("--decision-id", default=None)
    approve_language.add_argument("--notes", default="")
    approve_language.set_defaults(func=command_approve_language)

    active_question = subparsers.add_parser("set-active-question", help="set the current sequential question")
    active_question.add_argument("--question", required=True, choices=question_choices())
    active_question.add_argument("--defer-later", action="store_true")
    active_question.set_defaults(func=command_set_active_question)

    prepare_schemes = subparsers.add_parser("prepare-schemes", help="prepare A/B/C scheme files and Codex prompt")
    prepare_schemes.add_argument("--question", required=True, choices=question_choices())
    prepare_schemes.add_argument("--overwrite", action="store_true")
    prepare_schemes.add_argument("--force", action="store_true")
    prepare_schemes.set_defaults(func=command_prepare_schemes)

    approval_brief = subparsers.add_parser("create-approval-brief", help="create a user-facing A/B/C scheme approval brief")
    approval_brief.add_argument("--question", required=True, choices=question_choices())
    approval_brief.add_argument("--force", action="store_true")
    approval_brief.set_defaults(func=command_create_approval_brief)

    model_brief = subparsers.add_parser("create-model-confirmation-brief", help="create user-facing model confirmation options")
    model_brief.add_argument("--question", required=True, choices=question_choices())
    model_brief.add_argument("--scheme", choices=["A", "B", "C", "a", "b", "c"], default=None)
    model_brief.add_argument("--force", action="store_true")
    model_brief.set_defaults(func=command_create_model_confirmation_brief)

    approve_schemes = subparsers.add_parser("approve-schemes", help="record user approval for schemes")
    approve_schemes.add_argument("--question", required=True, choices=question_choices())
    approve_schemes.add_argument("--schemes", default="A,B,C")
    approve_schemes.add_argument("--decision-id", default=None)
    approve_schemes.add_argument("--notes", default="")
    approve_schemes.add_argument("--force", action="store_true")
    approve_schemes.set_defaults(func=command_approve_schemes)

    workorder = subparsers.add_parser("create-workorder", help="create a Claude Code workorder from template")
    workorder.add_argument("--question", required=True, choices=question_choices())
    workorder.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    workorder.add_argument("--workorder-id", default=None)
    workorder.add_argument("--out", type=Path, default=None)
    workorder.add_argument("--force", action="store_true")
    workorder.set_defaults(func=command_create_workorder)

    workorders = subparsers.add_parser("create-workorders", help="create Claude Code workorders for multiple schemes")
    workorders.add_argument("--question", required=True, choices=question_choices())
    workorders.add_argument("--schemes", default="A,B,C")
    workorders.add_argument("--force", action="store_true")
    workorders.set_defaults(func=command_create_workorders)

    claude_prompt = subparsers.add_parser("create-claude-prompt", help="create a copyable Claude Code execution prompt")
    claude_prompt.add_argument("--question", required=True, choices=question_choices())
    claude_prompt.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    claude_prompt.add_argument("--force", action="store_true")
    claude_prompt.set_defaults(func=command_create_claude_prompt)

    dispatch = subparsers.add_parser("dispatch-claude", help="send a prompt to Claude Code; auto opens a visible Terminal session")
    dispatch.add_argument("--question", required=True, choices=question_choices())
    dispatch.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    dispatch.add_argument("--prompt", type=Path, default=None)
    dispatch.add_argument("--revision", action="store_true", help="dispatch the latest revision prompt")
    dispatch.add_argument("--mode", choices=["auto", "terminal", "cli"], default="auto")
    dispatch.add_argument("--command", default="")
    dispatch.add_argument("--terminal-app", default="Terminal", choices=["Terminal", "iTerm"])
    dispatch.add_argument("--terminal-permission-mode", default="default", choices=["acceptEdits", "auto", "bypassPermissions", "default", "dontAsk", "plan"])
    dispatch.add_argument("--timeout", type=float, default=0.0)
    dispatch.add_argument("--watch", action="store_true")
    dispatch.add_argument("--watch-timeout", type=float, default=0.0)
    dispatch.add_argument("--interval", type=float, default=30.0)
    dispatch.add_argument("--require-standard-outputs", action="store_true")
    dispatch.set_defaults(func=command_dispatch_claude)

    monitor = subparsers.add_parser("open-claude-monitor", help="open a visible Terminal dashboard for Claude task status")
    monitor.add_argument("--question", required=True, choices=question_choices())
    monitor.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    monitor.add_argument("--interval", type=float, default=5.0)
    monitor.add_argument("--terminal-app", default="Terminal", choices=["Terminal", "iTerm"])
    monitor.add_argument("--target-os", default="auto", choices=["auto", "windows", "posix"])
    monitor.add_argument("--no-open", action="store_true", help="only create the monitor script")
    monitor.set_defaults(func=command_open_claude_monitor)

    vscode_tasks = subparsers.add_parser(
        "install-vscode-tasks",
        help="install VS Code integrated-terminal tasks for Claude dispatch and monitoring",
    )
    vscode_tasks.add_argument("--question", required=True, choices=question_choices())
    vscode_tasks.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    vscode_tasks.add_argument("--prompt", type=Path, default=None)
    vscode_tasks.add_argument("--revision", action="store_true", help="use the latest revision prompt")
    vscode_tasks.add_argument("--interval", type=float, default=5.0)
    vscode_tasks.add_argument("--target-os", default="auto", choices=["auto", "windows", "posix"])
    vscode_tasks.add_argument(
        "--terminal-permission-mode",
        default="default",
        choices=["acceptEdits", "auto", "bypassPermissions", "default", "dontAsk", "plan"],
    )
    vscode_tasks.set_defaults(func=command_install_vscode_tasks)

    ingest = subparsers.add_parser("ingest-claude-report", help="ingest a Claude Code completion report")
    ingest.add_argument("--question", required=True, choices=question_choices())
    ingest.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    ingest.add_argument("--report", type=Path, required=True)
    ingest.set_defaults(func=command_ingest_claude_report)

    check_claude = subparsers.add_parser("check-claude", help="check whether Claude Code has written a completion report")
    check_claude.add_argument("--question", required=True, choices=question_choices())
    check_claude.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    check_claude.add_argument("--ingest", action="store_true")
    check_claude.add_argument("--create-review", action="store_true")
    check_claude.add_argument("--require-standard-outputs", action="store_true")
    check_claude.add_argument("--fail-missing", action="store_true")
    check_claude.set_defaults(func=command_check_claude)

    watch_claude = subparsers.add_parser("watch-claude", help="poll for Claude Code completion and optionally ingest it")
    watch_claude.add_argument("--question", required=True, choices=question_choices())
    watch_claude.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    watch_claude.add_argument("--interval", type=float, default=30.0)
    watch_claude.add_argument("--timeout", type=float, default=0.0)
    watch_claude.add_argument("--once", action="store_true")
    watch_claude.add_argument("--ingest", action="store_true")
    watch_claude.add_argument("--create-review", action="store_true")
    watch_claude.add_argument("--require-standard-outputs", action="store_true")
    watch_claude.add_argument("--quiet", action="store_true")
    watch_claude.set_defaults(func=command_watch_claude)

    review = subparsers.add_parser("create-review", help="create a Codex review template for one scheme")
    review.add_argument("--question", required=True, choices=question_choices())
    review.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    review.set_defaults(func=command_create_review)

    mark_reviewed = subparsers.add_parser("mark-reviewed", help="record Codex review result")
    mark_reviewed.add_argument("--question", required=True, choices=question_choices())
    mark_reviewed.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    mark_reviewed.add_argument("--result", required=True, choices=["PASS", "REVISE", "BLOCKED", "pass", "revise", "blocked"])
    mark_reviewed.set_defaults(func=command_mark_reviewed)

    compare = subparsers.add_parser("compare-schemes", help="create scheme comparison template")
    compare.add_argument("--question", required=True, choices=question_choices())
    compare.set_defaults(func=command_compare_schemes)

    confirm = subparsers.add_parser("confirm-model", help="record final model confirmation")
    confirm.add_argument("--question", required=True, choices=question_choices())
    confirm.add_argument("--scheme", required=True, choices=["A", "B", "C", "a", "b", "c"])
    confirm.add_argument("--decision-id", default=None)
    confirm.add_argument("--notes", default="")
    confirm.add_argument("--force", action="store_true")
    confirm.set_defaults(func=command_confirm_model)

    figures = subparsers.add_parser("approve-figures", help="record user approval for final figures")
    figures.add_argument("--question", required=True, choices=question_choices())
    figures.add_argument("--figures", default="")
    figures.add_argument("--decision-id", default=None)
    figures.add_argument("--notes", default="")
    figures.add_argument("--allow-non-chinese", action="store_true")
    figures.set_defaults(func=command_approve_figures)

    paper_written = subparsers.add_parser("mark-paper-written", help="mark one question as written into the paper")
    paper_written.add_argument("--question", required=True, choices=question_choices())
    paper_written.set_defaults(func=command_mark_paper_written)

    write_question_paper = subparsers.add_parser("write-question-paper", help="write one confirmed question into LaTeX and compile")
    write_question_paper.add_argument("--question", required=True, choices=question_choices())
    write_question_paper.add_argument("--force", action="store_true")
    write_question_paper.add_argument("--compile", action=argparse.BooleanOptionalAction, default=True)
    write_question_paper.set_defaults(func=command_write_question_paper)

    finalize_summary = subparsers.add_parser("finalize-summary-paper", help="guard final abstract/evaluation writing until all questions are written")
    finalize_summary.add_argument("--force", action="store_true")
    finalize_summary.add_argument("--compile", action=argparse.BooleanOptionalAction, default=True)
    finalize_summary.set_defaults(func=command_finalize_summary_paper)

    paper_check = subparsers.add_parser("paper-check", help="check required paper files")
    paper_check.set_defaults(func=command_paper_check)

    layout_check = subparsers.add_parser("layout-check", help="check PDF layout gaps and expected figure/table pages")
    layout_check.add_argument("--pdf", type=Path, default=Path("07_paper/main.pdf"))
    layout_check.add_argument("--aux", type=Path, default=Path("07_paper/main.aux"))
    layout_check.add_argument("--max-internal-gap-ratio", type=float, default=0.24)
    layout_check.add_argument("--expect-label-page", action="append", default=[])
    layout_check.set_defaults(func=command_layout_check)

    readiness = subparsers.add_parser("readiness", help="print workflow readiness summary")
    readiness.set_defaults(func=command_readiness)

    latex = subparsers.add_parser("latex-check", help="compile the current LaTeX paper")
    latex.add_argument("--skip-layout-check", action="store_true")
    latex.add_argument("--max-internal-gap-ratio", type=float, default=0.24)
    latex.add_argument("--expect-label-page", action="append", default=[])
    latex.set_defaults(func=command_latex_check)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    locked_commands = {
        "init-state",
        "status",
        "import-problem",
        "scan-data",
        "approve-language",
        "set-active-question",
        "prepare-schemes",
        "create-approval-brief",
        "create-model-confirmation-brief",
        "approve-schemes",
        "create-workorder",
        "create-workorders",
        "create-claude-prompt",
        "ingest-claude-report",
        "check-claude",
        "create-review",
        "mark-reviewed",
        "compare-schemes",
        "confirm-model",
        "approve-figures",
        "mark-paper-written",
        "write-question-paper",
        "finalize-summary-paper",
        "layout-check",
        "readiness",
    }
    if args.command in locked_commands:
        with workflow_lock():
            args.func(args)
    else:
        args.func(args)


if __name__ == "__main__":
    main()
