"""Data scanning: walk a data directory and build a Markdown data dictionary."""

from __future__ import annotations

import re
from pathlib import Path

from ._util import now_iso, rel, write_text


def detect_question_ids(text: str) -> list[str]:
    from ._paths import QUESTIONS

    found = []
    for match in re.finditer(r"问题\s*([1-9])", text):
        qid = f"Q{match.group(1)}"
        if qid in QUESTIONS and qid not in found:
            found.append(qid)
    return found


def inspect_data_file(path: Path, sample_rows: int = 200, include_unsupported: bool = False) -> list[dict]:
    suffix = path.suffix.lower()
    rows: list[dict] = []
    try:
        if suffix in [".csv", ".txt"]:
            import pandas as pd

            df = pd.read_csv(path, nrows=sample_rows)
            total_rows = count_text_rows(path)
            rows.extend(column_rows(path, df, total_rows=total_rows, sheet=""))
        elif suffix in [".xlsx", ".xls"]:
            import pandas as pd

            xls = pd.ExcelFile(path)
            excel_rows = count_excel_rows(path) if suffix == ".xlsx" else {}
            for sheet in xls.sheet_names:
                df = pd.read_excel(path, sheet_name=sheet, nrows=sample_rows)
                rows.extend(column_rows(path, df, total_rows=excel_rows.get(sheet), sheet=sheet))
        elif suffix == ".json":
            import pandas as pd

            try:
                df = pd.read_json(path, lines=True, nrows=sample_rows)
            except TypeError:
                df = pd.read_json(path)
            rows.extend(column_rows(path, df.head(sample_rows), total_rows=None, sheet=""))
        elif include_unsupported:
            rows.append(file_row(path, "", "", "", f"unsupported file type {suffix}"))
    except Exception as exc:  # noqa: BLE001 - diagnostic tool, surface scan errors.
        rows.append(file_row(path, "", "", "", f"scan failed: {exc}"))
    return rows


def count_text_rows(path: Path):
    try:
        with path.open("rb") as f:
            return max(sum(1 for _ in f) - 1, 0)
    except OSError:
        return None


def count_excel_rows(path: Path) -> dict[str, int]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        return {sheet.title: max((sheet.max_row or 1) - 1, 0) for sheet in workbook.worksheets}
    except Exception:  # noqa: BLE001 - row counts are helpful but non-critical.
        return {}


def column_rows(path: Path, df, total_rows, sheet: str) -> list[dict]:
    rows = []
    for col in df.columns:
        series = df[col]
        rows.append(
            file_row(
                path,
                sheet,
                str(col),
                str(series.dtype),
                f"sample_non_null={int(series.notna().sum())}; total_rows={total_rows if total_rows is not None else 'unknown'}",
            )
        )
    if len(df.columns) == 0:
        rows.append(file_row(path, sheet, "", "", "empty or no columns detected"))
    return rows


def file_row(path: Path, sheet: str, column: str, dtype: str, note: str) -> dict:
    return {
        "file": rel(path),
        "sheet": sheet,
        "column": column,
        "dtype": dtype,
        "unit": "待 Codex 判断",
        "meaning": "待 Codex 根据题面判断",
        "quality": note,
    }


def write_data_dictionary(out_path: Path, rows: list[dict], data_dir: Path) -> None:
    lines = [
        "# Data Dictionary",
        "",
        f"- 数据目录：`{rel(data_dir)}`",
        f"- 扫描时间：{now_iso()}",
        "",
        "## 字段字典",
        "",
        "| 文件 | 工作表 | 字段 | 类型 | 单位 | 含义 | 数据质量 |",
        "|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        lines.append(
            f"| `{row['file']}` | {row['sheet'] or '-'} | {row['column'] or '-'} | {row['dtype'] or '-'} | {row['unit']} | {row['meaning']} | {row['quality']} |"
        )
    lines.extend(
        [
            "",
            "## 数据清洗规则",
            "",
            "- 待 Claude Code 根据 Codex 工单实现，Codex 负责审查。",
            "",
            "## 数据使用风险",
            "",
            "- 待 Codex 根据题面、字段和缺失异常情况判断。",
            "",
        ]
    )
    write_text(out_path, "\n".join(lines))


def data_dictionary_is_placeholder(path: Path) -> bool:
    """True if the file is the unfilled template (scan-data has not been run yet)."""
    if not path.exists():
        return True
    try:
        text = path.read_text(encoding="utf-8")
    except OSError:
        return True
    # The placeholder file contains "- 数据文件：待扫描" but no field rows.
    return "待扫描" in text and "## 字段字典" not in text
